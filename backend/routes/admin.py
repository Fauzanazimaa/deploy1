from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
from models import db, User, DataType, Task, Submission, ManualEntry, ExcelTemplate
from utils.excel import generate_template, parse_excel_structure
from storage import upload_file, download_file, delete_file, get_public_url, UPLOADS_BUCKET, TEMPLATES_BUCKET
from datetime import datetime
import io
import json
import uuid
import openpyxl

admin_bp = Blueprint('admin', __name__)


def require_admin():
    user_id = get_jwt_identity()
    user = User.query.get(int(user_id))
    if not user or user.role != 'admin':
        return None, jsonify({'error': 'Admin access required'}), 403
    return user, None, None


def _read_excel_headers(file_bytes: bytes) -> list:
    """Baca baris header pertama dari bytes Excel."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for cell in row:
            if cell is not None:
                headers.append(str(cell).strip())
    wb.close()
    return headers


# ─── Users ───────────────────────────────────────────────────────────────────

@admin_bp.route('/users', methods=['GET'])
@jwt_required()
def get_users():
    user, err, code = require_admin()
    if err:
        return err, code
    return jsonify([u.to_dict() for u in User.query.all()]), 200


@admin_bp.route('/users', methods=['POST'])
@jwt_required()
def create_user():
    user, err, code = require_admin()
    if err:
        return err, code

    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    for field in ['username', 'email', 'password', 'role']:
        if not data.get(field):
            return jsonify({'error': f'{field} is required'}), 400

    if data['role'] not in ['admin', 'contributor', 'viewer']:
        return jsonify({'error': 'Invalid role'}), 400
    if User.query.filter_by(username=data['username']).first():
        return jsonify({'error': 'Username already exists'}), 409
    if User.query.filter_by(email=data['email']).first():
        return jsonify({'error': 'Email already exists'}), 409

    new_user = User(
        username=data['username'],
        email=data['email'],
        password_hash=generate_password_hash(data['password']),
        role=data['role'],
        is_active=data.get('is_active', True),
    )
    db.session.add(new_user)
    db.session.commit()
    return jsonify(new_user.to_dict()), 201


@admin_bp.route('/users/<int:user_id>', methods=['PUT'])
@jwt_required()
def update_user(user_id):
    current_user, err, code = require_admin()
    if err:
        return err, code

    target = User.query.get_or_404(user_id)
    data = request.get_json()

    if 'username' in data:
        existing = User.query.filter_by(username=data['username']).first()
        if existing and existing.id != user_id:
            return jsonify({'error': 'Username already exists'}), 409
        target.username = data['username']
    if 'email' in data:
        existing = User.query.filter_by(email=data['email']).first()
        if existing and existing.id != user_id:
            return jsonify({'error': 'Email already exists'}), 409
        target.email = data['email']
    if 'role' in data:
        if data['role'] not in ['admin', 'contributor', 'viewer']:
            return jsonify({'error': 'Invalid role'}), 400
        target.role = data['role']
    if 'is_active' in data:
        target.is_active = data['is_active']
    if 'password' in data and data['password']:
        target.password_hash = generate_password_hash(data['password'])

    db.session.commit()
    return jsonify(target.to_dict()), 200


@admin_bp.route('/users/<int:user_id>', methods=['DELETE'])
@jwt_required()
def delete_user(user_id):
    current_user, err, code = require_admin()
    if err:
        return err, code

    target = User.query.get_or_404(user_id)
    if target.id == int(get_jwt_identity()):
        return jsonify({'error': 'Cannot delete yourself'}), 400

    try:
        Submission.query.filter_by(contributor_id=target.id).delete()
        tasks_assigned = Task.query.filter_by(assigned_to=target.id).all()
        for task in tasks_assigned:
            Submission.query.filter_by(task_id=task.id).delete()
            db.session.delete(task)
        Task.query.filter_by(assigned_by=target.id).update({'assigned_by': int(get_jwt_identity())})
        ManualEntry.query.filter_by(entered_by=target.id).delete()
        db.session.delete(target)
        db.session.commit()
        return jsonify({'message': 'User deleted'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete user: {str(e)}'}), 500


# ─── Data Types ───────────────────────────────────────────────────────────────

@admin_bp.route('/data-types', methods=['GET'])
@jwt_required()
def get_data_types():
    user, err, code = require_admin()
    if err:
        return err, code
    return jsonify([dt.to_dict() for dt in DataType.query.all()]), 200


@admin_bp.route('/data-types', methods=['POST'])
@jwt_required()
def create_data_type():
    user, err, code = require_admin()
    if err:
        return err, code

    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'error': 'Name is required'}), 400

    dt = DataType(
        name=data['name'],
        description=data.get('description', ''),
        fields_schema=json.dumps(data.get('fields_schema', [])),
        created_by=int(get_jwt_identity()),
    )
    db.session.add(dt)
    db.session.commit()
    return jsonify(dt.to_dict()), 201


@admin_bp.route('/data-types/<int:dt_id>', methods=['PUT'])
@jwt_required()
def update_data_type(dt_id):
    user, err, code = require_admin()
    if err:
        return err, code

    dt = DataType.query.get_or_404(dt_id)
    data = request.get_json()

    if 'name' in data:
        dt.name = data['name']
    if 'description' in data:
        dt.description = data['description']
    if 'fields_schema' in data:
        dt.fields_schema = json.dumps(data['fields_schema'])

    db.session.commit()
    return jsonify(dt.to_dict()), 200


@admin_bp.route('/data-types/<int:dt_id>', methods=['DELETE'])
@jwt_required()
def delete_data_type(dt_id):
    user, err, code = require_admin()
    if err:
        return err, code

    dt = DataType.query.get_or_404(dt_id)
    try:
        # 1. Hapus semua submission dari setiap task
        tasks = Task.query.filter_by(data_type_id=dt_id).all()
        for task in tasks:
            subs = Submission.query.filter_by(task_id=task.id).all()
            for sub in subs:
                try:
                    delete_file(UPLOADS_BUCKET(), sub.file_path)
                except Exception:
                    pass
            Submission.query.filter_by(task_id=task.id).delete()
            db.session.delete(task)

        # 2. Hapus template
        templates = ExcelTemplate.query.filter_by(data_type_id=dt_id).all()
        for tmpl in templates:
            try:
                delete_file(TEMPLATES_BUCKET(), tmpl.file_path)
            except Exception:
                pass
            db.session.delete(tmpl)

        # 3. Hapus manual entries
        ManualEntry.query.filter_by(data_type_id=dt_id).delete()

        # 4. Hapus dashboard widgets yang terkait
        from models import DashboardWidget
        DashboardWidget.query.filter_by(data_type_id=dt_id).delete()

        # 5. Baru hapus data type-nya
        db.session.delete(dt)
        db.session.commit()
        return jsonify({'message': 'Data type deleted'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete data type: {str(e)}'}), 500


# ─── Tasks ────────────────────────────────────────────────────────────────────

@admin_bp.route('/tasks', methods=['GET'])
@jwt_required()
def get_tasks():
    user, err, code = require_admin()
    if err:
        return err, code
    return jsonify([t.to_dict() for t in Task.query.all()]), 200


@admin_bp.route('/tasks', methods=['POST'])
@jwt_required()
def create_task():
    user, err, code = require_admin()
    if err:
        return err, code

    data = request.get_json()
    for field in ['title', 'data_type_id', 'assigned_to']:
        if not data.get(field):
            return jsonify({'error': f'{field} is required'}), 400

    assignee = User.query.get(data['assigned_to'])
    if not assignee or assignee.role != 'contributor':
        return jsonify({'error': 'Assigned user must be a contributor'}), 400

    deadline = None
    if data.get('deadline'):
        try:
            deadline = datetime.fromisoformat(data['deadline'])
        except ValueError:
            return jsonify({'error': 'Invalid deadline format'}), 400

    task = Task(
        title=data['title'],
        description=data.get('description', ''),
        data_type_id=data['data_type_id'],
        assigned_to=data['assigned_to'],
        assigned_by=int(get_jwt_identity()),
        deadline=deadline,
    )
    db.session.add(task)
    db.session.commit()
    return jsonify(task.to_dict()), 201


@admin_bp.route('/tasks/<int:task_id>', methods=['PUT'])
@jwt_required()
def update_task(task_id):
    user, err, code = require_admin()
    if err:
        return err, code

    task = Task.query.get_or_404(task_id)
    data = request.get_json()

    if 'title' in data:
        task.title = data['title']
    if 'description' in data:
        task.description = data['description']
    if 'status' in data:
        task.status = data['status']
    if 'deadline' in data:
        task.deadline = datetime.fromisoformat(data['deadline']) if data['deadline'] else None

    db.session.commit()
    return jsonify(task.to_dict()), 200


@admin_bp.route('/tasks/<int:task_id>', methods=['DELETE'])
@jwt_required()
def delete_task(task_id):
    user, err, code = require_admin()
    if err:
        return err, code

    task = Task.query.get_or_404(task_id)
    try:
        # Hapus file submission dari storage
        for sub in task.submissions:
            try:
                delete_file(UPLOADS_BUCKET(), sub.file_path)
            except Exception:
                pass  # lanjut meski file tidak ditemukan
        Submission.query.filter_by(task_id=task.id).delete()
        db.session.delete(task)
        db.session.commit()
        return jsonify({'message': 'Task deleted'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Gagal menghapus tugas: {str(e)}'}), 500


# ─── Templates ────────────────────────────────────────────────────────────────

@admin_bp.route('/templates', methods=['GET'])
@jwt_required()
def get_templates():
    user, err, code = require_admin()
    if err:
        return err, code
    return jsonify([t.to_dict() for t in ExcelTemplate.query.all()]), 200


@admin_bp.route('/templates', methods=['POST'])
@jwt_required()
def upload_template():
    user, err, code = require_admin()
    if err:
        return err, code

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    data_type_id = request.form.get('data_type_id')
    sync_schema  = request.form.get('sync_schema', 'false').lower() == 'true'

    if not data_type_id:
        return jsonify({'error': 'data_type_id is required'}), 400

    dt = DataType.query.get_or_404(int(data_type_id))

    # Enforce 1 template per data type — hapus lama jika ada
    existing = ExcelTemplate.query.filter_by(data_type_id=dt.id).first()
    if existing:
        delete_file(TEMPLATES_BUCKET(), existing.file_path)
        db.session.delete(existing)

    file_bytes = file.read()
    filename   = secure_filename(file.filename)
    storage_path = f"{uuid.uuid4()}_{filename}"

    upload_file(TEMPLATES_BUCKET(), storage_path, file_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Sync field schema dari struktur Excel jika diminta
    if sync_schema:
        try:
            parsed = parse_excel_structure(file_bytes)
            dt.fields_schema = json.dumps(parsed)
        except Exception:
            pass

    template = ExcelTemplate(
        data_type_id=dt.id,
        file_path=storage_path,
        original_filename=filename,
        created_by=int(get_jwt_identity()),
    )
    db.session.add(template)
    db.session.commit()
    return jsonify(template.to_dict()), 201


@admin_bp.route('/templates/parse', methods=['POST'])
@jwt_required()
def parse_template_structure():
    """Preview struktur Excel yang diupload tanpa menyimpan apapun."""
    user, err, code = require_admin()
    if err:
        return err, code

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    try:
        file_bytes = file.read()
        schema = parse_excel_structure(file_bytes)
        return jsonify({'schema': schema}), 200
    except Exception as e:
        return jsonify({'error': f'Gagal membaca struktur Excel: {str(e)}'}), 400


@admin_bp.route('/templates/generate', methods=['POST'])
@jwt_required()
def generate_excel_template():
    user, err, code = require_admin()
    if err:
        return err, code

    data  = request.get_json()
    dt_id = data.get('data_type_id')
    if not dt_id:
        return jsonify({'error': 'data_type_id is required'}), 400

    dt = DataType.query.get_or_404(dt_id)

    # Enforce 1 template per data type — hapus lama jika ada
    existing = ExcelTemplate.query.filter_by(data_type_id=dt.id).first()
    if existing:
        delete_file(TEMPLATES_BUCKET(), existing.file_path)
        db.session.delete(existing)

    # Generate ke buffer
    file_bytes   = generate_template(dt)
    storage_path = f"template_{dt.id}_{uuid.uuid4()}.xlsx"

    upload_file(TEMPLATES_BUCKET(), storage_path, file_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    template = ExcelTemplate(
        data_type_id=dt.id,
        file_path=storage_path,
        original_filename=f"{dt.name}_template.xlsx",
        created_by=int(get_jwt_identity()),
    )
    db.session.add(template)
    db.session.commit()
    return jsonify(template.to_dict()), 201


@admin_bp.route('/templates/<int:template_id>/download', methods=['GET'])
@jwt_required()
def download_template(template_id):
    template = ExcelTemplate.query.get_or_404(template_id)
    try:
        file_bytes = download_file(TEMPLATES_BUCKET(), template.file_path)
    except Exception:
        return jsonify({'error': 'File not found in storage'}), 404

    return send_file(
        io.BytesIO(file_bytes),
        as_attachment=True,
        download_name=template.original_filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@admin_bp.route('/templates/<int:template_id>', methods=['DELETE'])
@jwt_required()
def delete_template(template_id):
    user, err, code = require_admin()
    if err:
        return err, code

    template = ExcelTemplate.query.get_or_404(template_id)
    force = request.args.get('force', 'false').lower() == 'true'

    if not force:
        active_tasks = Task.query.filter(
            Task.data_type_id == template.data_type_id,
            Task.status.in_(['pending', 'submitted', 'revision'])
        ).all()

        if active_tasks:
            titles = [t.title for t in active_tasks[:3]]
            more   = len(active_tasks) - 3
            msg    = f"Template sedang digunakan oleh {len(active_tasks)} tugas aktif: {', '.join(titles)}"
            if more > 0:
                msg += f", dan {more} lainnya"
            return jsonify({'error': msg, 'active_tasks': len(active_tasks)}), 409

    try:
        delete_file(TEMPLATES_BUCKET(), template.file_path)
        db.session.delete(template)
        db.session.commit()
        return jsonify({'message': 'Template deleted'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete template: {str(e)}'}), 500


# ─── Submissions ──────────────────────────────────────────────────────────────

@admin_bp.route('/submissions', methods=['GET'])
@jwt_required()
def get_submissions():
    user, err, code = require_admin()
    if err:
        return err, code
    submissions = Submission.query.order_by(Submission.submitted_at.desc()).all()
    return jsonify([s.to_dict() for s in submissions]), 200


@admin_bp.route('/submissions/<int:sub_id>/approve', methods=['PUT'])
@jwt_required()
def approve_submission(sub_id):
    user, err, code = require_admin()
    if err:
        return err, code

    submission = Submission.query.get_or_404(sub_id)
    submission.status      = 'approved'
    submission.reviewed_at = datetime.utcnow()
    if submission.task:
        submission.task.status = 'approved'
    db.session.commit()
    return jsonify(submission.to_dict()), 200


@admin_bp.route('/submissions/<int:sub_id>/revision', methods=['PUT'])
@jwt_required()
def revision_submission(sub_id):
    user, err, code = require_admin()
    if err:
        return err, code

    data = request.get_json()
    if not data or not data.get('revision_notes'):
        return jsonify({'error': 'Revision notes are required'}), 400

    submission = Submission.query.get_or_404(sub_id)
    submission.status         = 'revision'
    submission.revision_notes = data['revision_notes']
    submission.reviewed_at    = datetime.utcnow()
    if submission.task:
        submission.task.status = 'revision'
    db.session.commit()
    return jsonify(submission.to_dict()), 200


@admin_bp.route('/submissions/<int:sub_id>/preview', methods=['GET'])
@jwt_required()
def preview_submission_by_id(sub_id):
    """Preview data submission langsung by ID, dengan schema yang benar."""
    user, err, code = require_admin()
    if err:
        return err, code

    submission = Submission.query.get_or_404(sub_id)

    # ── Preview submission dari form (JSON) ───────────────────────────────────
    if submission.source == 'form' and submission.form_data:
        try:
            import json as _json
            form_data = _json.loads(submission.form_data)
            dt = submission.task.data_type if submission.task else None
            schema_raw = dt.get_fields_schema() if dt else {}
            from utils.excel import normalize_schema, get_leaf_columns
            schema    = normalize_schema(schema_raw)
            first_col = schema.get('first_column', {})
            has_first = first_col.get('enabled', False)
            leaf_cols = get_leaf_columns(schema)

            # Bangun headers: [fc_label, col1, col2, ...]
            headers = []
            if has_first:
                headers.append(first_col.get('label', 'Baris'))
            headers += [f.get('label', f.get('name', f'Kol')) for f in leaf_cols]

            # Bangun group_headers untuk tampilan multi-level
            levels = schema.get('header_levels', [[]])
            group_levels = []
            if len(levels) > 1:
                for li in range(len(levels) - 1):
                    group_levels.append(levels[li])

            rows_out = []
            for row_obj in (form_data if isinstance(form_data, list) else []):
                row = {}
                if has_first:
                    row[first_col.get('label', 'Baris')] = row_obj.get('__row_label', '')
                for f in leaf_cols:
                    fname = f.get('name', '')
                    row[f.get('label', fname)] = row_obj.get(fname, '')
                rows_out.append(row)

            return jsonify({
                'rows': rows_out,
                'headers': headers,
                'group_levels': group_levels,
                'has_first_col': has_first,
                'first_col_label': first_col.get('label', '') if has_first else '',
                'total': len(rows_out),
                'source': 'form'
            }), 200
        except Exception as e:
            return jsonify({'error': f'Gagal membaca form data: {str(e)}'}), 400

    # ── Preview submission dari Excel ─────────────────────────────────────────
    if not submission.file_path:
        return jsonify({'error': 'Submission tidak memiliki file'}), 404

    try:
        file_bytes = download_file(UPLOADS_BUCKET(), submission.file_path)
    except Exception:
        return jsonify({'error': 'File tidak ditemukan di storage'}), 404

    try:
        # Gunakan fields_schema dari DataType jika ada, untuk parse yang benar
        dt = submission.task.data_type if submission.task else None
        schema_raw = dt.get_fields_schema() if dt else None

        if schema_raw:
            from utils.excel import normalize_schema, get_leaf_columns, read_excel_data
            schema    = normalize_schema(schema_raw)
            first_col = schema.get('first_column', {})
            has_first = first_col.get('enabled', False)
            leaf_cols = get_leaf_columns(schema)
            levels    = schema.get('header_levels', [[]])

            # Baca data dengan schema
            raw_rows = read_excel_data(file_bytes, schema_raw)

            # Konversi ke format dengan label kolom (bukan nama field)
            fc_key = None
            if has_first:
                from utils.excel import _to_field_name
                fc_key = _to_field_name(first_col.get('label', 'kolom_1'))

            headers = []
            if has_first:
                headers.append(first_col.get('label', 'Baris'))
            headers += [f.get('label', f.get('name', '')) for f in leaf_cols]

            rows_out = []
            for rd in raw_rows:
                row = {}
                if has_first and fc_key:
                    row[first_col.get('label', 'Baris')] = rd.get(fc_key, '')
                for f in leaf_cols:
                    fname = f.get('name', '')
                    flabel = f.get('label', fname)
                    val = rd.get(fname, '')
                    row[flabel] = '' if val is None else (val.isoformat() if hasattr(val, 'isoformat') else val)
                rows_out.append(row)

            # Group levels untuk header multi-level
            group_levels = []
            if len(levels) > 1:
                for li in range(len(levels) - 1):
                    group_levels.append(levels[li])

            return jsonify({
                'rows': rows_out,
                'headers': headers,
                'group_levels': group_levels,
                'has_first_col': has_first,
                'first_col_label': first_col.get('label', '') if has_first else '',
                'total': len(rows_out),
                'source': 'excel'
            }), 200

        else:
            # Fallback: baca sebagai plain Excel tanpa schema
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            rows_out = []
            headers_list = None
            for row in ws.iter_rows(values_only=True):
                if all(v is None for v in row):
                    continue
                if headers_list is None:
                    headers_list = [str(h).strip() if h is not None else f'Kolom_{i+1}'
                                    for i, h in enumerate(row)]
                    continue
                rd = {}
                for i, val in enumerate(row):
                    if i < len(headers_list):
                        rd[headers_list[i]] = '' if val is None else (val.isoformat() if hasattr(val, 'isoformat') else val)
                if any(v != '' for v in rd.values()):
                    rows_out.append(rd)
            wb.close()
            return jsonify({
                'rows': rows_out,
                'headers': headers_list or [],
                'group_levels': [],
                'has_first_col': False,
                'first_col_label': '',
                'total': len(rows_out),
                'source': 'excel'
            }), 200
    except Exception as e:
        return jsonify({'error': f'Gagal membaca file: {str(e)}'}), 400


@admin_bp.route('/submissions/preview', methods=['POST'])
@jwt_required()
def preview_submission():
    """Legacy: preview upload file langsung (masih dipakai untuk backward compat)."""
    user, err, code = require_admin()
    if err:
        return err, code

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    try:
        file_bytes = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows_out = []
        headers = None
        for row in ws.iter_rows(values_only=True):
            if all(v is None for v in row):
                continue
            if headers is None:
                headers = [str(h).strip() if h is not None else f'Kolom_{i+1}'
                           for i, h in enumerate(row)]
                continue
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    key = headers[i]
                    if val is None:
                        row_dict[key] = ''
                    elif hasattr(val, 'isoformat'):
                        row_dict[key] = val.isoformat()
                    else:
                        row_dict[key] = val
            if any(v != '' and v is not None for v in row_dict.values()):
                rows_out.append(row_dict)
        wb.close()
        return jsonify({'rows': rows_out, 'total': len(rows_out)}), 200
    except Exception as e:
        return jsonify({'error': f'Gagal membaca file: {str(e)}'}), 400


@admin_bp.route('/submissions/<int:sub_id>/download', methods=['GET'])
@jwt_required()
def download_submission(sub_id):
    user, err, code = require_admin()
    if err:
        return err, code

    submission = Submission.query.get_or_404(sub_id)

    # ── Jika submission dari form, konversi form_data ke Excel ────────────────
    if submission.source == 'form' and submission.form_data:
        try:
            from utils.excel import normalize_schema, _build_excel
            import json as _json

            form_data = _json.loads(submission.form_data)
            dt = submission.task.data_type if submission.task else None
            schema_raw = dt.get_fields_schema() if dt else {}
            schema = normalize_schema(schema_raw)

            # Bangun Excel dari form_data
            file_bytes = _form_data_to_excel(form_data, schema, dt.name if dt else 'Data')
            fname = f"submission_{sub_id}_{(dt.name if dt else 'data').replace(' ','_')}.xlsx"
            return send_file(
                io.BytesIO(file_bytes),
                as_attachment=True,
                download_name=fname,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        except Exception as e:
            return jsonify({'error': f'Gagal konversi form ke Excel: {str(e)}'}), 500

    # ── Submission dari upload Excel biasa ────────────────────────────────────
    try:
        file_bytes = download_file(UPLOADS_BUCKET(), submission.file_path)
    except Exception:
        return jsonify({'error': 'File not found in storage'}), 404

    return send_file(
        io.BytesIO(file_bytes),
        as_attachment=True,
        download_name=submission.file_path.split('_', 1)[-1] if '_' in submission.file_path else submission.file_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def _form_data_to_excel(form_data: dict, schema: dict, sheet_name: str) -> bytes:
    """Konversi form_data JSON ke file Excel sesuai schema."""
    from utils.excel import _build_excel, get_leaf_columns
    import openpyxl, io as _io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    levels    = schema.get('header_levels', [[]])
    first_col = schema.get('first_column', {})
    has_first = first_col.get('enabled', False)
    fc_label  = first_col.get('label', '')
    fc_rows   = first_col.get('default_rows', [])
    leaf_cols = levels[-1] if levels else []
    num_levels = len(levels)
    col_offset = 2 if has_first else 1

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    COLORS = ['1E3A5F', '2563EB', '3B82F6', '60A5FA']
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
    valign = Alignment(vertical='center')

    # ── Header rows ───────────────────────────────────────────────────────────
    for lvl_idx, level in enumerate(levels):
        row = lvl_idx + 1
        if has_first and lvl_idx == 0:
            cell = ws.cell(row=1, column=1)
            if num_levels > 1:
                ws.merge_cells(start_row=1, start_column=1, end_row=num_levels, end_column=1)
            cell.value = fc_label
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
            cell.alignment = halign
            cell.border = thin

        is_last = (lvl_idx == num_levels - 1)
        color = COLORS[min(lvl_idx, len(COLORS)-1)]
        hfont = Font(bold=True, color='FFFFFF', size=11 if lvl_idx==0 else 10)
        hfill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        if not is_last:
            col = col_offset
            for grp in level:
                span = grp.get('span', 1)
                cell = ws.cell(row=row, column=col)
                cell.value = grp.get('label', '')
                cell.font = hfont; cell.fill = hfill
                cell.alignment = halign; cell.border = thin
                if span > 1:
                    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
                    for cc in range(col, col+span):
                        ws.cell(row=row, column=cc).border = thin
                col += span
        else:
            for ci, field in enumerate(level):
                col = col_offset + ci
                label = field.get('label', field.get('name', f'Kol {ci+1}'))
                cell = ws.cell(row=row, column=col)
                cell.value = label
                cell.font = hfont; cell.fill = hfill
                cell.alignment = halign; cell.border = thin
                ws.column_dimensions[get_column_letter(col)].width = max(15, len(label)+5)

    if has_first:
        ws.column_dimensions['A'].width = max(20, len(fc_label)+5)
    for r in range(1, num_levels+1):
        ws.row_dimensions[r].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    # form_data format: {row_key: {field_name: value}} atau [{row: ..., fields: {...}}]
    # Kita simpan sebagai list of dicts: [{fc_row_label, field1, field2, ...}]
    data_start = num_levels + 1
    rows_data = form_data if isinstance(form_data, list) else []
    n_rows = max(len(rows_data), len(fc_rows), 1)

    fc_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    fc_font = Font(bold=True, size=10)

    for ri in range(n_rows):
        row = data_start + ri
        row_data = rows_data[ri] if ri < len(rows_data) else {}
        if has_first:
            cell = ws.cell(row=row, column=1)
            cell.value = fc_rows[ri] if ri < len(fc_rows) else row_data.get('__row_label', '')
            cell.border = thin; cell.alignment = valign
            cell.fill = fc_fill; cell.font = fc_font
        for ci, field in enumerate(leaf_cols):
            fname = field.get('name', '')
            cell = ws.cell(row=row, column=col_offset+ci)
            cell.value = row_data.get(fname, '')
            cell.border = thin; cell.alignment = valign

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@admin_bp.route('/submissions/<int:sub_id>', methods=['DELETE'])
@jwt_required()
def delete_submission(sub_id):
    """Hapus satu submission beserta file-nya dari storage."""
    user, err, code = require_admin()
    if err:
        return err, code

    submission = Submission.query.get_or_404(sub_id)
    try:
        # Reset status tugas ke pending agar kontributor bisa kirim ulang
        if submission.task and submission.task.status in ('submitted', 'revision', 'approved'):
            submission.task.status = 'pending'
        # Hapus file dari storage
        try:
            delete_file(UPLOADS_BUCKET(), submission.file_path)
        except Exception:
            pass
        db.session.delete(submission)
        db.session.commit()
        return jsonify({'message': 'Submission deleted'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Gagal menghapus: {str(e)}'}), 500




# ─── Manual Entries ───────────────────────────────────────────────────────────

@admin_bp.route('/manual-entries', methods=['GET'])
@jwt_required()
def get_manual_entries():
    user, err, code = require_admin()
    if err:
        return err, code
    entries = ManualEntry.query.order_by(ManualEntry.created_at.desc()).all()
    return jsonify([e.to_dict() for e in entries]), 200


@admin_bp.route('/manual-entries', methods=['POST'])
@jwt_required()
def create_manual_entry():
    user, err, code = require_admin()
    if err:
        return err, code

    data = request.get_json()
    if not data or not data.get('data_type_id'):
        return jsonify({'error': 'data_type_id is required'}), 400

    entry = ManualEntry(
        data_type_id=data['data_type_id'],
        data=json.dumps(data.get('data', {})),
        entered_by=int(get_jwt_identity()),
    )
    db.session.add(entry)
    db.session.commit()
    return jsonify(entry.to_dict()), 201


@admin_bp.route('/manual-entries/<int:entry_id>', methods=['DELETE'])
@jwt_required()
def delete_manual_entry(entry_id):
    user, err, code = require_admin()
    if err:
        return err, code

    entry = ManualEntry.query.get_or_404(entry_id)
    db.session.delete(entry)
    db.session.commit()
    return jsonify({'message': 'Entry deleted'}), 200


# ─── Dashboard Stats ──────────────────────────────────────────────────────────

@admin_bp.route('/dashboard/stats', methods=['GET'])
@jwt_required()
def dashboard_stats():
    user, err, code = require_admin()
    if err:
        return err, code

    recent_submissions = Submission.query.order_by(Submission.submitted_at.desc()).limit(5).all()

    return jsonify({
        'total_users':          User.query.count(),
        'total_contributors':   User.query.filter_by(role='contributor').count(),
        'total_viewers':        User.query.filter_by(role='viewer').count(),
        'total_tasks':          Task.query.count(),
        'task_status': {
            'pending':   Task.query.filter_by(status='pending').count(),
            'submitted': Task.query.filter_by(status='submitted').count(),
            'approved':  Task.query.filter_by(status='approved').count(),
            'revision':  Task.query.filter_by(status='revision').count(),
        },
        'total_submissions':      Submission.query.count(),
        'pending_verifications':  Submission.query.filter_by(status='pending').count(),
        'total_data_types':       DataType.query.count(),
        'total_manual_entries':   ManualEntry.query.count(),
        'recent_submissions':     [s.to_dict() for s in recent_submissions],
    }), 200
