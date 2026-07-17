from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, User, Task, Submission, ManualEntry, DataType
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
from datetime import datetime

viewer_bp = Blueprint('viewer', __name__)


def get_current_user():
    return User.query.get(int(get_jwt_identity()))


@viewer_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def viewer_dashboard():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    data_types = DataType.query.all()
    data_type_stats = []
    for dt in data_types:
        approved_count = db.session.query(func.count(Task.id)).filter(
            Task.data_type_id == dt.id, Task.status == 'approved'
        ).scalar()
        manual_count = ManualEntry.query.filter_by(data_type_id=dt.id).count()
        data_type_stats.append({
            'id': dt.id, 'name': dt.name,
            'approved_tasks': approved_count,
            'manual_entries': manual_count,
            'total': approved_count + manual_count,
        })

    return jsonify({
        'approved_tasks':      Task.query.filter_by(status='approved').count(),
        'total_data_types':    DataType.query.count(),
        'total_manual_entries': ManualEntry.query.count(),
        'approved_submissions': Submission.query.filter_by(status='approved').count(),
        'data_type_stats':     data_type_stats,
    }), 200


@viewer_bp.route('/data', methods=['GET'])
@jwt_required()
def get_approved_data():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    data_type_id = request.args.get('data_type_id', type=int)

    query = ManualEntry.query
    if data_type_id:
        query = query.filter_by(data_type_id=data_type_id)
    manual_entries = query.order_by(ManualEntry.created_at.desc()).all()

    sub_query = Submission.query.filter_by(status='approved')
    if data_type_id:
        sub_query = sub_query.join(Task).filter(Task.data_type_id == data_type_id)
    approved_submissions = sub_query.order_by(Submission.reviewed_at.desc()).all()

    return jsonify({
        'manual_entries':      [e.to_dict() for e in manual_entries],
        'approved_submissions': [s.to_dict() for s in approved_submissions],
    }), 200


@viewer_bp.route('/data-types', methods=['GET'])
@jwt_required()
def get_data_types():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify([dt.to_dict() for dt in DataType.query.all()]), 200


@viewer_bp.route('/data/export', methods=['POST'])
@jwt_required()
def export_data():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    data_type_id = data.get('data_type_id')
    if not data_type_id:
        return jsonify({'error': 'data_type_id required'}), 400

    data_type = DataType.query.get_or_404(data_type_id)
    fields = data_type.get_fields_schema()

    manual_entries = (ManualEntry.query
                      .filter_by(data_type_id=data_type_id)
                      .order_by(ManualEntry.created_at.desc()).all())
    approved_submissions = (Submission.query
                            .filter_by(status='approved')
                            .join(Task).filter(Task.data_type_id == data_type_id)
                            .order_by(Submission.reviewed_at.desc()).all())

    # Build Excel in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = data_type.name[:31]

    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    headers = (['No', 'Sumber']
               + [f.get('label', f.get('name', f'Field {i+1}')) for i, f in enumerate(fields)]
               + ['Tanggal Input'])

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    row_idx = 2
    for idx, entry in enumerate(manual_entries, 1):
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value='Entri Manual')
        entry_data = entry.get_data()
        for ci, field in enumerate(fields, 3):
            ws.cell(row=row_idx, column=ci, value=entry_data.get(field.get('name', ''), '-'))
        ws.cell(row=row_idx, column=len(headers), value=entry.created_at.strftime('%Y-%m-%d %H:%M'))
        row_idx += 1

    for sub in approved_submissions:
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=f'Upload ({sub.contributor.username})')
        ws.cell(row=row_idx, column=3, value='Data tersedia di file upload asli')
        ws.cell(row=row_idx, column=len(headers),
                value=sub.reviewed_at.strftime('%Y-%m-%d') if sub.reviewed_at else '-')
        row_idx += 1

    for col in ws.columns:
        width = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"export_{data_type.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
