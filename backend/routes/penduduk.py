"""
Routes untuk Indikator Penduduk Kabupaten Sijunjung.
Admin bisa input/update data. Viewer bisa baca tanpa login.
"""
from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, User
from datetime import datetime

penduduk_bp = Blueprint('penduduk', __name__)

# ── Helper ────────────────────────────────────────────────────────────────────

def _require_admin():
    uid = get_jwt_identity()
    u = User.query.get(int(uid))
    if not u or u.role != 'admin':
        return None, jsonify({'error': 'Admin access required'}), 403
    return u, None, None


def _init_tables():
    """Buat tabel jika belum ada, dan migrate kolom baru bila perlu."""
    from sqlalchemy import text
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS penduduk_jk (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tahun INTEGER NOT NULL,
            laki_laki INTEGER DEFAULT 0,
            perempuan INTEGER DEFAULT 0,
            updated_at TEXT,
            UNIQUE(tahun)
        )
    """))
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS penduduk_umur (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tahun INTEGER NOT NULL,
            kelompok_umur TEXT NOT NULL,
            laki_laki INTEGER DEFAULT 0,
            perempuan INTEGER DEFAULT 0,
            updated_at TEXT,
            UNIQUE(tahun, kelompok_umur)
        )
    """))
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS penduduk_kec (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tahun INTEGER NOT NULL,
            kecamatan TEXT NOT NULL,
            laki_laki INTEGER DEFAULT 0,
            perempuan INTEGER DEFAULT 0,
            updated_at TEXT,
            UNIQUE(tahun, kecamatan)
        )
    """))
    db.session.commit()

    # Migrate: tambah kolom laki_laki/perempuan kalau tabel lama hanya punya jumlah
    for tbl in ('penduduk_umur', 'penduduk_kec'):
        cols = [r[1] for r in db.session.execute(text(f"PRAGMA table_info({tbl})")).fetchall()]
        if 'laki_laki' not in cols:
            db.session.execute(text(f"ALTER TABLE {tbl} ADD COLUMN laki_laki INTEGER DEFAULT 0"))
        if 'perempuan' not in cols:
            db.session.execute(text(f"ALTER TABLE {tbl} ADD COLUMN perempuan INTEGER DEFAULT 0"))
        # Isi laki_laki dari jumlah lama agar data tidak hilang
        if 'jumlah' in cols:
            db.session.execute(text(
                f"UPDATE {tbl} SET laki_laki=COALESCE(jumlah,0) WHERE laki_laki=0 AND jumlah>0"
            ))
    db.session.commit()


# ── PUBLIC: baca data ─────────────────────────────────────────────────────────

@penduduk_bp.route('/public/penduduk/jk', methods=['GET'])
def get_penduduk_jk():
    _init_tables()
    from sqlalchemy import text
    rows = db.session.execute(text(
        "SELECT tahun, laki_laki, perempuan FROM penduduk_jk ORDER BY tahun"
    )).fetchall()
    return jsonify([{
        'tahun': r[0], 'laki_laki': r[1] or 0, 'perempuan': r[2] or 0,
        'total': (r[1] or 0) + (r[2] or 0)
    } for r in rows]), 200


@penduduk_bp.route('/public/penduduk/umur', methods=['GET'])
def get_penduduk_umur():
    _init_tables()
    from sqlalchemy import text
    tahun_filter = request.args.get('tahun')
    if tahun_filter:
        rows = db.session.execute(text(
            "SELECT tahun, kelompok_umur, laki_laki, perempuan FROM penduduk_umur WHERE tahun=:t ORDER BY kelompok_umur"
        ), {'t': int(tahun_filter)}).fetchall()
    else:
        rows = db.session.execute(text(
            "SELECT tahun, kelompok_umur, laki_laki, perempuan FROM penduduk_umur ORDER BY tahun, kelompok_umur"
        )).fetchall()

    tahun_list = db.session.execute(text(
        "SELECT DISTINCT tahun FROM penduduk_umur ORDER BY tahun"
    )).fetchall()

    return jsonify({
        'data': [{
            'tahun': r[0], 'kelompok_umur': r[1],
            'laki_laki': r[2] or 0, 'perempuan': r[3] or 0,
            'total': (r[2] or 0) + (r[3] or 0)
        } for r in rows],
        'tahun_tersedia': [r[0] for r in tahun_list]
    }), 200


@penduduk_bp.route('/public/penduduk/kecamatan', methods=['GET'])
def get_penduduk_kec():
    _init_tables()
    from sqlalchemy import text
    rows = db.session.execute(text(
        "SELECT tahun, kecamatan, laki_laki, perempuan FROM penduduk_kec ORDER BY tahun, kecamatan"
    )).fetchall()

    tahun_list = db.session.execute(text(
        "SELECT DISTINCT tahun FROM penduduk_kec ORDER BY tahun"
    )).fetchall()

    return jsonify({
        'data': [{
            'tahun': r[0], 'kecamatan': r[1],
            'laki_laki': r[2] or 0, 'perempuan': r[3] or 0,
            'total': (r[2] or 0) + (r[3] or 0)
        } for r in rows],
        'tahun_tersedia': [r[0] for r in tahun_list]
    }), 200


# ── ADMIN: input/update data ──────────────────────────────────────────────────

@penduduk_bp.route('/admin/penduduk/jk', methods=['GET'])
@jwt_required()
def admin_get_jk():
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    from sqlalchemy import text
    rows = db.session.execute(text(
        "SELECT id, tahun, laki_laki, perempuan, updated_at FROM penduduk_jk ORDER BY tahun"
    )).fetchall()
    return jsonify([{
        'id': r[0], 'tahun': r[1], 'laki_laki': r[2] or 0,
        'perempuan': r[3] or 0, 'updated_at': r[4]
    } for r in rows]), 200


@penduduk_bp.route('/admin/penduduk/jk', methods=['POST'])
@jwt_required()
def admin_upsert_jk():
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    from sqlalchemy import text
    data = request.get_json()
    if not data or not data.get('tahun'):
        return jsonify({'error': 'tahun wajib diisi'}), 400
    now = datetime.utcnow().isoformat()
    db.session.execute(text("""
        INSERT INTO penduduk_jk (tahun, laki_laki, perempuan, updated_at)
        VALUES (:t, :lk, :pr, :now)
        ON CONFLICT(tahun) DO UPDATE SET
            laki_laki=excluded.laki_laki, perempuan=excluded.perempuan,
            updated_at=excluded.updated_at
    """), {'t': data['tahun'], 'lk': data.get('laki_laki', 0),
           'pr': data.get('perempuan', 0), 'now': now})
    db.session.commit()
    return jsonify({'message': 'Tersimpan'}), 200


@penduduk_bp.route('/admin/penduduk/jk/<int:tahun>', methods=['DELETE'])
@jwt_required()
def admin_delete_jk(tahun):
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    from sqlalchemy import text
    db.session.execute(text("DELETE FROM penduduk_jk WHERE tahun=:t"), {'t': tahun})
    db.session.commit()
    return jsonify({'message': 'Dihapus'}), 200


@penduduk_bp.route('/admin/penduduk/umur', methods=['GET'])
@jwt_required()
def admin_get_umur():
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    from sqlalchemy import text
    rows = db.session.execute(text(
        "SELECT id, tahun, kelompok_umur, laki_laki, perempuan, updated_at FROM penduduk_umur ORDER BY tahun, kelompok_umur"
    )).fetchall()
    return jsonify([{
        'id': r[0], 'tahun': r[1], 'kelompok_umur': r[2],
        'laki_laki': r[3] or 0, 'perempuan': r[4] or 0, 'updated_at': r[5]
    } for r in rows]), 200


@penduduk_bp.route('/admin/penduduk/umur', methods=['POST'])
@jwt_required()
def admin_upsert_umur():
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    from sqlalchemy import text
    data = request.get_json()
    if not data or not data.get('tahun') or not data.get('kelompok_umur'):
        return jsonify({'error': 'tahun dan kelompok_umur wajib diisi'}), 400
    now = datetime.utcnow().isoformat()
    rows = data if isinstance(data, list) else [data]
    for row in rows:
        db.session.execute(text("""
            INSERT INTO penduduk_umur (tahun, kelompok_umur, laki_laki, perempuan, updated_at)
            VALUES (:t, :ku, :lk, :pr, :now)
            ON CONFLICT(tahun, kelompok_umur) DO UPDATE SET
                laki_laki=excluded.laki_laki, perempuan=excluded.perempuan,
                updated_at=excluded.updated_at
        """), {'t': row['tahun'], 'ku': row['kelompok_umur'],
               'lk': row.get('laki_laki', 0), 'pr': row.get('perempuan', 0), 'now': now})
    db.session.commit()
    return jsonify({'message': f'{len(rows)} baris tersimpan'}), 200


@penduduk_bp.route('/admin/penduduk/umur/<int:row_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_umur(row_id):
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    from sqlalchemy import text
    db.session.execute(text("DELETE FROM penduduk_umur WHERE id=:id"), {'id': row_id})
    db.session.commit()
    return jsonify({'message': 'Dihapus'}), 200


@penduduk_bp.route('/admin/penduduk/kecamatan', methods=['GET'])
@jwt_required()
def admin_get_kec():
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    from sqlalchemy import text
    rows = db.session.execute(text(
        "SELECT id, tahun, kecamatan, laki_laki, perempuan, updated_at FROM penduduk_kec ORDER BY tahun, kecamatan"
    )).fetchall()
    return jsonify([{
        'id': r[0], 'tahun': r[1], 'kecamatan': r[2],
        'laki_laki': r[3] or 0, 'perempuan': r[4] or 0, 'updated_at': r[5]
    } for r in rows]), 200


@penduduk_bp.route('/admin/penduduk/kecamatan', methods=['POST'])
@jwt_required()
def admin_upsert_kec():
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    from sqlalchemy import text
    data = request.get_json()
    rows = data if isinstance(data, list) else [data]
    now = datetime.utcnow().isoformat()
    for row in rows:
        if not row.get('tahun') or not row.get('kecamatan'):
            continue
        db.session.execute(text("""
            INSERT INTO penduduk_kec (tahun, kecamatan, laki_laki, perempuan, updated_at)
            VALUES (:t, :kec, :lk, :pr, :now)
            ON CONFLICT(tahun, kecamatan) DO UPDATE SET
                laki_laki=excluded.laki_laki, perempuan=excluded.perempuan,
                updated_at=excluded.updated_at
        """), {'t': row['tahun'], 'kec': row['kecamatan'],
               'lk': row.get('laki_laki', 0), 'pr': row.get('perempuan', 0), 'now': now})
    db.session.commit()
    return jsonify({'message': f'{len(rows)} baris tersimpan'}), 200


@penduduk_bp.route('/admin/penduduk/kecamatan/<int:row_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_kec(row_id):
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    from sqlalchemy import text
    db.session.execute(text("DELETE FROM penduduk_kec WHERE id=:id"), {'id': row_id})
    db.session.commit()
    return jsonify({'message': 'Dihapus'}), 200


# ── TEMPLATE & UPLOAD ─────────────────────────────────────────────────────────

@penduduk_bp.route('/admin/penduduk/umur/template', methods=['GET'])
@jwt_required()
def download_template_umur():
    """Download template Excel untuk kelompok umur."""
    u, err, code = _require_admin()
    if err: return err, code
    import os, io
    from flask import send_file, current_app
    base = os.path.join(os.path.dirname(__file__), '..', 'templates_excel')
    path = os.path.abspath(os.path.join(base, 'template_penduduk_umur.xlsx'))
    if not os.path.exists(path):
        return jsonify({'error': 'Template tidak ditemukan'}), 404
    return send_file(path, as_attachment=True,
                     download_name='template_penduduk_umur.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@penduduk_bp.route('/admin/penduduk/umur/upload', methods=['POST'])
@jwt_required()
def upload_umur():
    """Upload Excel kelompok umur. Kolom: Tahun, Kelompok Umur, Jumlah."""
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    if 'file' not in request.files:
        return jsonify({'error': 'File tidak ditemukan'}), 400
    f = request.files['file']
    import openpyxl, io
    from sqlalchemy import text
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        now = datetime.utcnow().isoformat()
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None: continue
            try:
                tahun = int(row[0])
                ku    = str(row[1]).strip() if row[1] else None
                jml   = int(row[2] or 0)
            except: continue
            if not ku: continue
            db.session.execute(text("""
                INSERT INTO penduduk_umur (tahun, kelompok_umur, laki_laki, perempuan, updated_at)
                VALUES (:t, :ku, :jml, 0, :now)
                ON CONFLICT(tahun, kelompok_umur) DO UPDATE SET
                    laki_laki=excluded.laki_laki,
                    updated_at=excluded.updated_at
            """), {'t': tahun, 'ku': ku, 'jml': jml, 'now': now})
            count += 1
        db.session.commit()
        return jsonify({'message': f'{count} baris berhasil diimport'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@penduduk_bp.route('/admin/penduduk/umur/tahun/<int:tahun>', methods=['DELETE'])
@jwt_required()
def admin_delete_umur_tahun(tahun):
    """Hapus semua data umur untuk 1 tahun."""
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    from sqlalchemy import text
    db.session.execute(text("DELETE FROM penduduk_umur WHERE tahun=:t"), {'t': tahun})
    db.session.commit()
    return jsonify({'message': f'Data tahun {tahun} dihapus'}), 200


@penduduk_bp.route('/admin/penduduk/kecamatan/template', methods=['GET'])
@jwt_required()
def download_template_kec():
    """Download template Excel untuk kecamatan."""
    u, err, code = _require_admin()
    if err: return err, code
    import os
    from flask import send_file
    base = os.path.join(os.path.dirname(__file__), '..', 'templates_excel')
    path = os.path.abspath(os.path.join(base, 'template_penduduk_kecamatan.xlsx'))
    if not os.path.exists(path):
        return jsonify({'error': 'Template tidak ditemukan'}), 404
    return send_file(path, as_attachment=True,
                     download_name='template_penduduk_kecamatan.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@penduduk_bp.route('/admin/penduduk/kecamatan/upload', methods=['POST'])
@jwt_required()
def upload_kec():
    """Upload Excel kecamatan. Kolom: Tahun, Kecamatan, Jumlah."""
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    if 'file' not in request.files:
        return jsonify({'error': 'File tidak ditemukan'}), 400
    f = request.files['file']
    import openpyxl, io
    from sqlalchemy import text
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        now = datetime.utcnow().isoformat()
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None: continue
            try:
                tahun = int(row[0])
                kec   = str(row[1]).strip() if row[1] else None
                jml   = int(row[2] or 0)
            except: continue
            if not kec: continue
            db.session.execute(text("""
                INSERT INTO penduduk_kec (tahun, kecamatan, laki_laki, perempuan, updated_at)
                VALUES (:t, :kec, :jml, 0, :now)
                ON CONFLICT(tahun, kecamatan) DO UPDATE SET
                    laki_laki=excluded.laki_laki,
                    updated_at=excluded.updated_at
            """), {'t': tahun, 'kec': kec, 'jml': jml, 'now': now})
            count += 1
        db.session.commit()
        return jsonify({'message': f'{count} baris berhasil diimport'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@penduduk_bp.route('/admin/penduduk/kecamatan/tahun/<int:tahun>', methods=['DELETE'])
@jwt_required()
def admin_delete_kec_tahun(tahun):
    """Hapus semua data kecamatan untuk 1 tahun."""
    u, err, code = _require_admin()
    if err: return err, code
    _init_tables()
    from sqlalchemy import text
    db.session.execute(text("DELETE FROM penduduk_kec WHERE tahun=:t"), {'t': tahun})
    db.session.commit()
    return jsonify({'message': f'Data tahun {tahun} dihapus'}), 200
