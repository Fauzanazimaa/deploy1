from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from werkzeug.utils import secure_filename
from models import db, User, Task, Submission, ExcelTemplate
from storage import upload_file, download_file, UPLOADS_BUCKET, TEMPLATES_BUCKET
from utils.excel import normalize_schema, get_leaf_columns
from utils.excel_parser import parse_excel_to_grid
from datetime import datetime
import io
import uuid
import json as _json
import openpyxl

contributor_bp = Blueprint('contributor', __name__)


def require_contributor():
    user_id = get_jwt_identity()
    user = User.query.get(int(user_id))
    if not user or user.role != 'contributor':
        return None, jsonify({'error': 'Contributor access required'}), 403
    return user, None, None


ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@contributor_bp.route('/tasks', methods=['GET'])
@jwt_required()
def get_my_tasks():
    user, err, code = require_contributor()
    if err:
        return err, code

    tasks = Task.query.filter_by(assigned_to=user.id).order_by(Task.created_at.desc()).all()
    result = []
    for t in tasks:
        task_dict = t.to_dict()
        latest_sub = (Submission.query
                      .filter_by(task_id=t.id, contributor_id=user.id)
                      .order_by(Submission.submitted_at.desc())
                      .first())
        task_dict['latest_submission'] = latest_sub.to_dict() if latest_sub else None
        # Sertakan fields_schema dari DataType agar frontend bisa render form
        if t.data_type:
            task_dict['fields_schema'] = t.data_type.get_fields_schema()
        result.append(task_dict)
    return jsonify(result), 200


@contributor_bp.route('/tasks/<int:task_id>', methods=['GET'])
@jwt_required()
def get_task(task_id):
    user, err, code = require_contributor()
    if err:
        return err, code
    task = Task.query.filter_by(id=task_id, assigned_to=user.id).first_or_404()
    return jsonify(task.to_dict()), 200


@contributor_bp.route('/templates/<int:data_type_id>', methods=['GET'])
@jwt_required()
def download_template(data_type_id):
    user_id = get_jwt_identity()
    user = User.query.get(int(user_id))
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    template = (ExcelTemplate.query
                .filter_by(data_type_id=data_type_id)
                .order_by(ExcelTemplate.created_at.desc())
                .first())
    if not template:
        return jsonify({'error': 'No template found for this data type'}), 404

    try:
        file_bytes = download_file(TEMPLATES_BUCKET(), template.file_path)
    except Exception:
        return jsonify({'error': 'Template file not found in storage'}), 404

    return send_file(
        io.BytesIO(file_bytes),
        as_attachment=True,
        download_name=template.original_filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@contributor_bp.route('/tasks/<int:task_id>/submit', methods=['POST'])
@jwt_required()
def submit_task(task_id):
    user, err, code = require_contributor()
    if err:
        return err, code

    task = Task.query.filter_by(id=task_id, assigned_to=user.id).first_or_404()

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'Only Excel files (.xlsx, .xls) or CSV are allowed'}), 400

    file_bytes   = file.read()
    filename     = secure_filename(file.filename)
    storage_path = f"{uuid.uuid4()}_{filename}"

    upload_file(UPLOADS_BUCKET(), storage_path, file_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Hapus semua submission lama untuk task ini agar tidak double
    from storage import delete_file as _del
    old_subs = Submission.query.filter_by(task_id=task.id, contributor_id=user.id).all()
    for old in old_subs:
        if old.file_path:
            try:
                _del(UPLOADS_BUCKET(), old.file_path)
            except Exception:
                pass
        db.session.delete(old)

    submission = Submission(
        task_id=task.id,
        contributor_id=user.id,
        file_path=storage_path,
        source='excel',
        status='pending',
    )
    db.session.add(submission)
    task.status = 'submitted'
    db.session.commit()

    return jsonify(submission.to_dict()), 201


@contributor_bp.route('/tasks/<int:task_id>/template-grid', methods=['GET'])
@jwt_required()
def get_template_grid(task_id):
    """
    Baca template Excel untuk task ini dan kembalikan sebagai grid 2D.
    Response:
      {
        "headers": [  // header rows (tidak bisa diubah)
          [{"value": str, "rowspan": int, "colspan": int, "is_header": true}, ...]
        ],
        "rows": [     // baris data; setiap sel punya "locked" (true jika terisi dari template)
          [{"value": str, "locked": bool}, ...]
        ],
        "num_header_rows": int,
        "num_data_cols": int,
        "has_first_col": bool,
        "first_col_label": str
      }
    """
    user_id = get_jwt_identity()
    user = User.query.get(int(user_id))
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    task = Task.query.filter_by(id=task_id, assigned_to=user.id).first_or_404()

    template = (ExcelTemplate.query
                .filter_by(data_type_id=task.data_type_id)
                .order_by(ExcelTemplate.created_at.desc())
                .first())

    # ── Jika tidak ada template, fallback ke fields_schema ───────────────────
    if not template:
        if not task.data_type:
            return jsonify({'error': 'Template tidak tersedia'}), 404
        schema = normalize_schema(task.data_type.get_fields_schema())
        return _schema_to_grid(schema), 200

    try:
        file_bytes = download_file(TEMPLATES_BUCKET(), template.file_path)
    except Exception:
        # Fallback ke schema
        if task.data_type:
            schema = normalize_schema(task.data_type.get_fields_schema())
            return _schema_to_grid(schema), 200
        return jsonify({'error': 'File template tidak ditemukan'}), 404

    return parse_excel_to_grid(file_bytes), 200


def _excel_to_grid(file_bytes: bytes):
    """
    Baca Excel template → grid JSON untuk direct input.
    Header dibaca PERSIS dari template — tidak ada manipulasi struktur.
    Baris data: sel yang kosong di template = bisa diedit, sel berisi = terkunci.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # ── Merge map ─────────────────────────────────────────────────────────────
    merge_info  = {}   # (r,c) top-left → {rowspan, colspan}
    merged_skip = set()
    for mr in ws.merged_cells.ranges:
        rs = mr.max_row - mr.min_row + 1
        cs = mr.max_col - mr.min_col + 1
        merge_info[(mr.min_row, mr.min_col)] = {'rowspan': rs, 'colspan': cs}
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    merged_skip.add((r, c))

    # ── Deteksi jumlah baris header ───────────────────────────────────────────
    # Hitung dari merge vertikal yang dimulai di baris 1
    num_header_rows = 1
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row > num_header_rows:
            num_header_rows = mr.max_row
        if num_header_rows >= 5:
            break

    def cell_val(r, c):
        v = ws.cell(r, c).value
        if v is None:
            return ''
        if hasattr(v, 'isoformat'):
            return v.isoformat()
        return str(v).strip()

    # ── Header rows — baca PERSIS dari template, tidak diubah ────────────────
    HEADER_COLORS = ['#1e3a5f', '#2563eb', '#3b82f6', '#60a5fa']
    header_rows = []
    for r in range(1, num_header_rows + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            if (r, c) in merged_skip:
                continue
            mi = merge_info.get((r, c), {'rowspan': 1, 'colspan': 1})
            # Warna: kolom yang di-rowspan penuh = level 0, sisanya ikut level baris
            color_idx = min(r - 1, len(HEADER_COLORS) - 1)
            if mi['rowspan'] >= num_header_rows and num_header_rows > 1:
                color_idx = 0
            row_cells.append({
                'value':   cell_val(r, c),
                'rowspan': mi['rowspan'],
                'colspan': mi['colspan'],
                'bg':      HEADER_COLORS[color_idx],
            })
        header_rows.append(row_cells)

    # ── Deteksi first_column ──────────────────────────────────────────────────
    # has_first_col = True jika kolom 1 di header di-merge secara vertikal
    # (artinya kolom 1 = deskriptor baris, bukan nilai data)
    has_first_col  = False
    first_col_label = ''

    if num_header_rows > 1:
        mi = merge_info.get((1, 1))
        if mi and mi['rowspan'] >= num_header_rows:
            has_first_col   = True
            first_col_label = cell_val(1, 1)
    else:
        # Single-header: cek apakah baris data di kolom 1 sudah terisi dari template
        # Jika ya, anggap kolom 1 adalah label baris (first_col)
        data_start_check = 2
        sample_vals = [cell_val(r, 1) for r in range(data_start_check,
                       min(data_start_check + 10, max_row + 1))]
        filled = [v for v in sample_vals if v.strip()]
        if filled:
            has_first_col   = True
            first_col_label = cell_val(1, 1)  # label kolom pertama dari header

    # ── Baris data ────────────────────────────────────────────────────────────
    data_start   = num_header_rows + 1
    col_start    = 1  # selalu mulai dari kolom 1 — tidak ada penggeseran
    num_data_cols = max_col  # total kolom data termasuk first_col jika ada

    # Temukan baris data terakhir yang ada isinya (supaya tidak baca baris kosong footer)
    last_valid = data_start - 1
    for r in range(data_start, max_row + 1):
        row_any = any(cell_val(r, c).strip() for c in range(1, max_col + 1))
        if row_any:
            last_valid = r

    # Baca semua baris data — sel berisi dari template = terkunci,
    # sel kosong = bisa diedit kontributor
    data_rows = []
    for r in range(data_start, last_valid + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            v = cell_val(r, c)
            row_cells.append({'value': v, 'locked': bool(v.strip())})
        data_rows.append(row_cells)

    # Pastikan minimal 1 baris tersedia untuk input
    if not data_rows:
        data_rows.append([{'value': '', 'locked': False} for _ in range(max_col)])

    wb.close()

    # num_data_cols = jumlah kolom data (di luar first_col jika ada)
    actual_data_cols = max_col - (1 if has_first_col else 0)

    return jsonify({
        'headers':         header_rows,
        'rows':            data_rows,
        'num_header_rows': num_header_rows,
        'num_data_cols':   actual_data_cols,
        'has_first_col':   has_first_col,
        'first_col_label': first_col_label,
        'total_cols':      max_col,
    })


def _schema_to_grid(schema: dict):
    """Fallback: bangun grid dari fields_schema (tanpa file Excel)."""
    levels = schema.get('header_levels', [[]])
    first_col = schema.get('first_column', {})
    has_first = first_col.get('enabled', False)
    fc_label = first_col.get('label', '')
    fc_rows = first_col.get('default_rows', [])
    leaf_cols = get_leaf_columns(schema)
    num_levels = len(levels)
    HEADER_COLORS = ['#1e3a5f', '#2563eb', '#3b82f6', '#60a5fa']

    header_rows = []
    for li, level in enumerate(levels):
        is_last = (li == num_levels - 1)
        row_cells = []
        if has_first and li == 0:
            row_cells.append({
                'value': fc_label,
                'rowspan': num_levels,
                'colspan': 1,
                'bg': HEADER_COLORS[0],
            })
        if not is_last:
            for grp in level:
                row_cells.append({
                    'value': grp.get('label', ''),
                    'rowspan': 1,
                    'colspan': grp.get('span', 1),
                    'bg': HEADER_COLORS[min(li, len(HEADER_COLORS)-1)],
                })
        else:
            for f in level:
                row_cells.append({
                    'value': f.get('label', f.get('name', '')),
                    'rowspan': 1,
                    'colspan': 1,
                    'bg': HEADER_COLORS[min(li, len(HEADER_COLORS)-1)],
                })
        header_rows.append(row_cells)

    # Hanya gunakan baris yang sudah ada di default_rows (dari template)
    # Jika tidak ada default_rows, tampilkan 1 baris kosong saja
    n_rows = len(fc_rows) if has_first else max(len(fc_rows), 1)
    data_rows = []
    for ri in range(n_rows):
        row = []
        if has_first:
            val = fc_rows[ri] if ri < len(fc_rows) else ''
            row.append({'value': val, 'locked': bool(val.strip())})
        for _ in leaf_cols:
            row.append({'value': '', 'locked': False})
        data_rows.append(row)

    if not data_rows:
        empty_row = []
        if has_first:
            empty_row.append({'value': '', 'locked': False})
        for _ in leaf_cols:
            empty_row.append({'value': '', 'locked': False})
        data_rows.append(empty_row)

    return jsonify({
        'headers': header_rows,
        'rows': data_rows,
        'num_header_rows': num_levels,
        'num_data_cols': len(leaf_cols),
        'has_first_col': has_first,
        'first_col_label': fc_label,
        'total_cols': (1 if has_first else 0) + len(leaf_cols),
    })


@contributor_bp.route('/tasks/<int:task_id>/submit-form', methods=['POST'])
@jwt_required()
def submit_task_form(task_id):
    """Submit data via form (JSON), bukan file upload."""
    user, err, code = require_contributor()
    if err:
        return err, code

    task = Task.query.filter_by(id=task_id, assigned_to=user.id).first_or_404()

    if task.status not in ('pending', 'revision'):
        return jsonify({'error': f'Tugas berstatus {task.status}, tidak bisa disubmit'}), 400

    data = request.get_json()
    if not data or 'form_data' not in data:
        return jsonify({'error': 'form_data diperlukan'}), 400

    # Hapus semua submission lama untuk task ini agar tidak double
    old_subs = Submission.query.filter_by(task_id=task.id, contributor_id=user.id).all()
    for old in old_subs:
        db.session.delete(old)

    submission = Submission(
        task_id=task.id,
        contributor_id=user.id,
        file_path=None,
        source='form',
        form_data=_json.dumps(data['form_data']),
        status='pending',
    )
    db.session.add(submission)
    task.status = 'submitted'
    db.session.commit()

    return jsonify(submission.to_dict()), 201


@contributor_bp.route('/submissions', methods=['GET'])
@jwt_required()
def get_my_submissions():
    user, err, code = require_contributor()
    if err:
        return err, code
    submissions = (Submission.query
                   .filter_by(contributor_id=user.id)
                   .order_by(Submission.submitted_at.desc())
                   .all())
    return jsonify([s.to_dict() for s in submissions]), 200


@contributor_bp.route('/dashboard/stats', methods=['GET'])
@jwt_required()
def dashboard_stats():
    user, err, code = require_contributor()
    if err:
        return err, code

    recent_tasks = (Task.query
                    .filter_by(assigned_to=user.id)
                    .order_by(Task.created_at.desc())
                    .limit(5).all())

    return jsonify({
        'total_tasks':        Task.query.filter_by(assigned_to=user.id).count(),
        'pending_tasks':      Task.query.filter_by(assigned_to=user.id, status='pending').count(),
        'submitted_tasks':    Task.query.filter_by(assigned_to=user.id, status='submitted').count(),
        'approved_tasks':     Task.query.filter_by(assigned_to=user.id, status='approved').count(),
        'revision_tasks':     Task.query.filter_by(assigned_to=user.id, status='revision').count(),
        'total_submissions':  Submission.query.filter_by(contributor_id=user.id).count(),
        'pending_submissions': Submission.query.filter_by(contributor_id=user.id, status='pending').count(),
        'recent_tasks':       [t.to_dict() for t in recent_tasks],
    }), 200


@contributor_bp.route('/submissions/<int:sub_id>/preview', methods=['GET'])
@jwt_required()
def preview_submission(sub_id):
    user, err, code = require_contributor()
    if err: return err, code
    
    submission = Submission.query.filter_by(id=sub_id, contributor_id=user.id).first_or_404()
    
    # ── Form submission → konversi ke format grid 2D dengan template overlay ──
    if submission.source == 'form' and submission.form_data:
        try:
            import json as _json
            form_data = _json.loads(submission.form_data)
            rows_data = form_data if isinstance(form_data, list) else []
            dt = submission.task.data_type if submission.task else None
            
            template = None
            if dt:
                from models import ExcelTemplate as _ET
                template = (_ET.query
                            .filter_by(data_type_id=dt.id)
                            .order_by(_ET.created_at.desc())
                            .first())
            
            grid_rows = []
            if template:
                try:
                    from utils.excel_parser import parse_excel_to_preview_grid, _build_merge_index, detect_header_rows
                    tmpl_bytes = download_file(TEMPLATES_BUCKET(), template.file_path)
                    
                    _wb = openpyxl.load_workbook(io.BytesIO(tmpl_bytes), data_only=True)
                    _ws = _wb.active
                    _max_col = _ws.max_column or 1
                    _num_header_rows = detect_header_rows(_ws)
                    _top_left, _non_tl = _build_merge_index(_ws)
                    
                    _has_first_col = False
                    _mi1 = _top_left.get((1, 1))
                    if _mi1 and _mi1['rowspan'] >= _num_header_rows and _num_header_rows > 1:
                        _has_first_col = True
                    elif _num_header_rows == 1:
                        _data_start = _num_header_rows + 1
                        _fc_vals = [str(_ws.cell(_r2, 1).value or '').strip()
                                    for _r2 in range(_data_start, min(_data_start + 5, _ws.max_row + 1))
                                    if _ws.cell(_r2, 1).value]
                        _has_first_col = bool(_fc_vals)
                    _wb.close()

                    preview_res = parse_excel_to_preview_grid(tmpl_bytes)
                    grid_rows = preview_res['grid']
                    _num_data_cols = _max_col - (1 if _has_first_col else 0)

                    for ri, row_obj in enumerate(rows_data):
                        grid_idx = _num_header_rows + ri
                        if grid_idx >= len(grid_rows):
                            _row_cells = []
                            if _has_first_col:
                                _row_cells.append({
                                    'value': str(row_obj.get('__row_label', '') or ''),
                                    'rowspan': 1, 'colspan': 1,
                                    'is_header': False, 'is_first_col': True,
                                    'is_merged_child': False
                                })
                            for _i in range(_num_data_cols):
                                _v = row_obj.get(f'__col_{_i}', row_obj.get(f'col_{_i}', ''))
                                _row_cells.append({
                                    'value': '' if _v is None else str(_v),
                                    'rowspan': 1, 'colspan': 1,
                                    'is_header': False,
                                    'is_merged_child': False
                                })
                            grid_rows.append(_row_cells)
                        else:
                            if _has_first_col and len(grid_rows[grid_idx]) > 0:
                                if '__row_label' in row_obj:
                                    grid_rows[grid_idx][0]['value'] = str(row_obj.get('__row_label') or '')
                                grid_rows[grid_idx][0]['is_first_col'] = True
                            
                            col_start = 1 if _has_first_col else 0
                            for _i in range(_num_data_cols):
                                cell_idx = col_start + _i
                                if cell_idx < len(grid_rows[grid_idx]):
                                    _v = row_obj.get(f'__col_{_i}', row_obj.get(f'col_{_i}', ''))
                                    grid_rows[grid_idx][cell_idx]['value'] = '' if _v is None else str(_v)
                    
                    return jsonify({
                        'grid': grid_rows,
                        'num_header_rows': _num_header_rows,
                        'total_cols': _max_col,
                        'source': 'form'
                    }), 200
                except Exception as e:
                    print(f"Error in contributor preview_submission: {str(e)}")
                    pass

            # Fallback schema
            schema_raw = dt.get_fields_schema() if dt else {}
            from utils.excel import normalize_schema, get_leaf_columns
            schema    = normalize_schema(schema_raw)
            first_col = schema.get('first_column', {})
            has_first = first_col.get('enabled', False)
            leaf_cols = get_leaf_columns(schema)
            levels    = schema.get('header_levels', [[]])
            num_levels = len(levels)

            for li, level in enumerate(levels):
                is_last = (li == num_levels - 1)
                row_cells = []
                if has_first and li == 0:
                    row_cells.append({
                        'value': first_col.get('label', 'Baris'),
                        'rowspan': num_levels, 'colspan': 1,
                        'is_header': True, 'level': 0
                    })
                if not is_last:
                    for grp in level:
                        row_cells.append({
                            'value': grp.get('label', ''),
                            'rowspan': 1, 'colspan': grp.get('span', 1),
                            'is_header': True, 'level': li
                        })
                else:
                    for f in level:
                        row_cells.append({
                            'value': f.get('label', f.get('name', '')),
                            'rowspan': 1, 'colspan': 1,
                            'is_header': True, 'level': li
                        })
                grid_rows.append(row_cells)

            num_leaf = len(leaf_cols)
            for row_obj in rows_data:
                row_cells = []
                if has_first:
                    row_cells.append({
                        'value': str(row_obj.get('__row_label', '') or ''),
                        'rowspan': 1, 'colspan': 1,
                        'is_header': False, 'is_first_col': True
                    })
                for _i, f in enumerate(leaf_cols):
                    fname = f.get('name', '')
                    v = row_obj.get(f'__col_{_i}', row_obj.get(f'col_{_i}', row_obj.get(fname, '')))
                    row_cells.append({
                        'value': '' if v is None else str(v),
                        'rowspan': 1, 'colspan': 1,
                        'is_header': False
                    })
                grid_rows.append(row_cells)

            return jsonify({
                'grid': grid_rows,
                'num_header_rows': num_levels,
                'total_cols': (1 if has_first else 0) + num_leaf,
                'source': 'form'
            }), 200
        except Exception as e:
            return jsonify({'error': f'Gagal membaca form data: {str(e)}'}), 400

    # ── Excel submission → baca file mentah sebagai grid 2D ──────────────────
    if not submission.file_path:
        return jsonify({'error': 'Submission tidak memiliki file'}), 404

    try:
        file_bytes = download_file(UPLOADS_BUCKET(), submission.file_path)
    except Exception:
        return jsonify({'error': 'File tidak ditemukan di storage'}), 404

    try:
        from utils.excel_parser import parse_excel_to_preview_grid, read_submission_data
        result = parse_excel_to_preview_grid(file_bytes)
        data_rows = read_submission_data(file_bytes)
        result['rows']  = data_rows
        result['total'] = len(data_rows)
        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': f'Gagal membaca file: {str(e)}'}), 400


@contributor_bp.route('/submissions/<int:sub_id>/download', methods=['GET'])
@jwt_required()
def download_submission(sub_id):
    user, err, code = require_contributor()
    if err: return err, code

    submission = Submission.query.filter_by(id=sub_id, contributor_id=user.id).first_or_404()

    if submission.source == 'form' and submission.form_data:
        try:
            import json as _json
            from routes.admin import _form_data_to_excel_from_template, _form_data_to_excel

            form_data = _json.loads(submission.form_data)
            dt = submission.task.data_type if submission.task else None

            template = None
            if dt:
                from models import ExcelTemplate as _ET2
                template = (_ET2.query
                            .filter_by(data_type_id=dt.id)
                            .order_by(_ET2.created_at.desc())
                            .first())

            if template:
                try:
                    tmpl_bytes = download_file(TEMPLATES_BUCKET(), template.file_path)
                    file_bytes = _form_data_to_excel_from_template(form_data, tmpl_bytes, dt.name if dt else 'Data')
                except Exception:
                    schema_raw = dt.get_fields_schema() if dt else {}
                    from utils.excel import normalize_schema as _ns
                    schema = _ns(schema_raw)
                    file_bytes = _form_data_to_excel(form_data, schema, dt.name if dt else 'Data')
            else:
                schema_raw = dt.get_fields_schema() if dt else {}
                from utils.excel import normalize_schema as _ns
                schema = _ns(schema_raw)
                file_bytes = _form_data_to_excel(form_data, schema, dt.name if dt else 'Data')

            fname = f"submission_{sub_id}_{(dt.name if dt else 'data').replace(' ','_')}.xlsx"
            return send_file(
                io.BytesIO(file_bytes),
                as_attachment=True,
                download_name=fname,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        except Exception as e:
            return jsonify({'error': f'Gagal konversi form ke Excel: {str(e)}'}), 500

    try:
        file_bytes = download_file(UPLOADS_BUCKET(), submission.file_path)
    except Exception:
        return jsonify({'error': 'File not found in storage'}), 404

    return send_file(
        io.BytesIO(file_bytes),
        as_attachment=True,
        download_name=submission.file_path.split('_', 1)[-1] if '_' in submission.file_path else submission.file_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ─── Assignment Letters (Surat Tugas) ──────────────────────────────────────

@contributor_bp.route('/assignment-letters', methods=['GET'])
@jwt_required()
def get_my_assignment_letters():
    user, err, code = require_contributor()
    if err:
        return err, code

    from models import Task, AssignmentLetter

    # Find all tasks assigned to this contributor
    my_tasks = Task.query.filter_by(assigned_to=user.id).all()
    
    # Extract unique task titles
    unique_titles = list(set([t.title for t in my_tasks if t.title]))
    
    if not unique_titles:
        return jsonify([]), 200

    # Retrieve all assignment letters for these titles
    letters = AssignmentLetter.query.filter(AssignmentLetter.task_title.in_(unique_titles)).all()
    
    return jsonify([l.to_dict() for l in letters]), 200


@contributor_bp.route('/assignment-letters/<int:letter_id>/download', methods=['GET'])
@jwt_required()
def download_my_assignment_letter(letter_id):
    user, err, code = require_contributor()
    if err:
        return err, code

    from models import Task, AssignmentLetter
    from storage import download_file, LETTERS_BUCKET

    letter = AssignmentLetter.query.get_or_404(letter_id)
    
    # Security check: verify contributor is assigned to at least one task with this title
    has_access = Task.query.filter_by(assigned_to=user.id, title=letter.task_title).first() is not None
    if not has_access:
        return jsonify({'error': 'Access denied'}), 403

    try:
        file_bytes = download_file(LETTERS_BUCKET(), letter.file_path)
    except Exception:
        return jsonify({'error': 'File not found in storage'}), 404

    return send_file(
        io.BytesIO(file_bytes),
        as_attachment=True,
        download_name=letter.original_filename,
        mimetype='application/pdf'
    )

