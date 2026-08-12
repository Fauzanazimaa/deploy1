"""
excel_parser.py — Parser Excel generik yang mempertahankan struktur asli.

PRINSIP:
- Baca worksheet APA ADANYA dari metadata openpyxl.
- Tidak menebak struktur, tidak meratakan header, tidak menghapus baris/kolom kosong.
- Semua merged range dibaca dari ws.merged_cells.ranges (metadata, bukan nilai sel).
- Output konsisten: grid 2D dengan rowspan/colspan persis seperti file asli.

PUBLIC API:
  parse_excel_to_grid(file_bytes)          → dict  (untuk preview / direct input)
  read_submission_data(file_bytes)         → list  (untuk simpan data submission)
  detect_header_rows(ws)                   → int   (jumlah baris header)
"""

import io
import openpyxl


# ── Helpers ───────────────────────────────────────────────────────────────────

def _cell_str(cell) -> str:
    """Konversi nilai cell ke string; kosong jika None."""
    v = cell.value
    if v is None:
        return ''
    if hasattr(v, 'isoformat'):   # datetime / date
        return v.isoformat()
    return str(v)


def _get_actual_max_col(ws) -> int:
    """
    Hitung max_column yang sebenarnya dengan mengabaikan kolom di bagian kanan
    yang sepenuhnya kosong (tidak memiliki nilai apapun di seluruh baris).
    """
    max_col = ws.max_column or 1
    actual_max = 1
    for c in range(1, max_col + 1):
        col_has_value = False
        for r in range(1, (ws.max_row or 1) + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != '':
                col_has_value = True
                break
        if col_has_value:
            actual_max = c
    return actual_max


def _build_merge_index(ws):
    """
    Bangun dua struktur dari ws.merged_cells.ranges:

    top_left  : dict[(r,c)] → {rowspan, colspan, max_r, max_c}
                Hanya sel paling atas-kiri dari setiap merged range.

    non_tl    : set of (r,c)
                Semua sel yang merupakan bagian dari merged range
                KECUALI sel top-left.  Sel-sel ini harus di-skip saat render.
    """
    top_left = {}
    non_tl   = set()
    for mr in ws.merged_cells.ranges:
        rs = mr.max_row  - mr.min_row + 1
        cs = mr.max_col  - mr.min_col + 1
        top_left[(mr.min_row, mr.min_col)] = {
            'rowspan': rs,
            'colspan': cs,
            'max_r':   mr.max_row,
            'max_c':   mr.max_col,
        }
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    non_tl.add((r, c))
    return top_left, non_tl


def detect_header_rows(ws) -> int:
    """
    Tentukan jumlah baris header berdasarkan merged ranges yang dimulai
    di baris 1.  Ini adalah heuristik yang paling aman:
    - Jika ada sel di baris 1 yang di-merge secara vertikal hingga baris N,
      maka baris 1 s/d N dianggap header.
    - Jika tidak ada merge vertikal dari baris 1, header = 1 baris.
    - Cap: maksimal 10 baris header (hindari false positive dari data area).

    Perbedaan dengan kode lama:
    - Kode lama mengambil merge dari MANA SAJA di sheet, sehingga merge di
      area data bisa salah meningkatkan jumlah header.
    - Fungsi ini hanya melihat merge yang MIN_ROW == 1.
    """
    max_h = 1
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row > max_h:
            max_h = mr.max_row
            if max_h >= 10:
                break
    return max_h


# ── Core parser ───────────────────────────────────────────────────────────────

def parse_excel_to_grid(file_bytes: bytes) -> dict:
    """
    Baca file Excel dan kembalikan representasi grid 2D yang mempertahankan
    SEMUA aspek struktur asli:
      - Setiap sel dikembalikan dengan rowspan/colspan sesuai merged_cells.ranges.
      - Sel yang merupakan bagian dari merge (non-top-left) di-skip — frontend
        harus menggunakan rowspan/colspan untuk merender sel induk.
      - Baris kosong di luar area header TIDAK dihapus (dipertahankan).
      - Nilai sel tidak diubah/dinormalisasi selain konversi ke string.

    Returns:
      {
        'headers':         list[list[cell_obj]],   baris-baris header
        'rows':            list[list[cell_obj]],   baris-baris data
        'num_header_rows': int,
        'total_cols':      int,
        'total_rows':      int,   jumlah baris data
        'has_first_col':   bool,
        'first_col_label': str,
      }

    cell_obj header:  { value, rowspan, colspan, bg }
    cell_obj data:    { value, rowspan, colspan, locked }

    'locked' = True  jika sel terisi dari file upload (sel pre-filled dari template).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    max_row = ws.max_row or 1
    max_col = _get_actual_max_col(ws)

    top_left, non_tl = _build_merge_index(ws)
    num_header_rows  = detect_header_rows(ws)

    HEADER_COLORS = ['#1e3a5f', '#2563eb', '#3b82f6', '#60a5fa']

    # ── Header rows ───────────────────────────────────────────────────────────
    header_rows = []
    for r in range(1, num_header_rows + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            if (r, c) in non_tl:
                row_cells.append({
                    'value':   '',
                    'rowspan': 1,
                    'colspan': 1,
                    'is_merged_child': True,
                })
                continue
            cell = ws.cell(r, c)
            mi   = top_left.get((r, c), {'rowspan': 1, 'colspan': 1})

            # Pilih warna berdasarkan level baris; kolom yang span penuh = level 0
            color_idx = min(r - 1, len(HEADER_COLORS) - 1)
            if mi['rowspan'] >= num_header_rows and num_header_rows > 1:
                color_idx = 0

            row_cells.append({
                'value':   _cell_str(cell),
                'rowspan': mi['rowspan'],
                'colspan': mi['colspan'],
                'bg':      HEADER_COLORS[color_idx],
                'is_merged_child': False,
            })
        header_rows.append(row_cells)

    # ── Deteksi has_first_col ─────────────────────────────────────────────────
    # Kolom 1 dianggap first_col jika dalam area header ia di-merge vertikal
    # yang mencakup SEMUA baris header (rowspan == num_header_rows).
    has_first_col  = False
    first_col_label = ''

    mi1 = top_left.get((1, 1))
    if mi1 and mi1['rowspan'] >= num_header_rows and num_header_rows > 1:
        has_first_col   = True
        first_col_label = _cell_str(ws.cell(1, 1))
    elif num_header_rows == 1:
        # Single-header: scan seluruh baris data kolom 1 (bukan hanya 10 baris)
        data_start_check = num_header_rows + 1
        any_filled = False
        for r in range(data_start_check, max_row + 1):
            if _cell_str(ws.cell(r, 1)).strip():
                any_filled = True
                break
        if any_filled:
            has_first_col   = True
            first_col_label = _cell_str(ws.cell(1, 1))

    # ── Data rows ─────────────────────────────────────────────────────────────
    # PENTING: Semua baris data dipertahankan, termasuk baris kosong.
    # Baris yang sepenuhnya kosong TETAP diikutsertakan untuk menjaga
    # struktur asli (spacer rows, subtotal separators, dst.).
    data_start = num_header_rows + 1
    data_rows  = []

    for r in range(data_start, max_row + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            if (r, c) in non_tl:
                row_cells.append({
                    'value':   '',
                    'rowspan': 1,
                    'colspan': 1,
                    'locked':  True,
                    'is_merged_child': True,
                })
                continue
            cell = ws.cell(r, c)
            mi   = top_left.get((r, c), {'rowspan': 1, 'colspan': 1})
            val  = _cell_str(cell)
            row_cells.append({
                'value':   val,
                'rowspan': mi['rowspan'],
                'colspan': mi['colspan'],
                'locked':  bool(val.strip()),   # pre-filled = terkunci untuk direct input
                'is_merged_child': False,
            })
        data_rows.append(row_cells)

    # Pastikan minimal 1 baris kosong tersedia untuk input
    if not data_rows:
        empty_row = [{'value': '', 'rowspan': 1, 'colspan': 1, 'locked': False, 'is_merged_child': False}
                     for _ in range(max_col)]
        data_rows.append(empty_row)

    wb.close()

    num_data_cols = max_col - (1 if has_first_col else 0)

    return {
        'headers':         header_rows,
        'rows':            data_rows,
        'num_header_rows': num_header_rows,
        'num_data_cols':   num_data_cols,
        'has_first_col':   has_first_col,
        'first_col_label': first_col_label,
        'total_cols':      max_col,
        'total_rows':      len(data_rows),
    }


def parse_excel_to_preview_grid(file_bytes: bytes) -> dict:
    """
    Baca file Excel sebagai grid 2D untuk preview (admin, viewer).
    Mempertahankan SEMUA sel termasuk baris kosong.
    Tidak membedakan header/data — semua dikembalikan dalam satu 'grid'.

    cell_obj: { value, rowspan, colspan, is_header, bg }

    Returns:
      {
        'grid':            list[list[cell_obj]],
        'num_header_rows': int,
        'total_cols':      int,
        'source':          'excel',
      }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    max_row = ws.max_row or 1
    max_col = _get_actual_max_col(ws)

    top_left, non_tl = _build_merge_index(ws)
    num_header_rows  = detect_header_rows(ws)

    HEADER_COLORS = ['#1e3a5f', '#2563eb', '#3b82f6', '#60a5fa']
    grid_rows = []

    for r in range(1, max_row + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            if (r, c) in non_tl:
                row_cells.append({
                    'value':     '',
                    'rowspan':   1,
                    'colspan':   1,
                    'is_header': (r <= num_header_rows),
                    'bg':        None,
                    'is_merged_child': True,
                })
                continue
            cell      = ws.cell(r, c)
            mi        = top_left.get((r, c), {'rowspan': 1, 'colspan': 1})
            val       = _cell_str(cell)
            is_header = (r <= num_header_rows)

            bg = None
            if is_header:
                color_idx = min(r - 1, len(HEADER_COLORS) - 1)
                if mi['rowspan'] >= num_header_rows and num_header_rows > 1:
                    color_idx = 0
                bg = HEADER_COLORS[color_idx]

            row_cells.append({
                'value':     val,
                'rowspan':   mi['rowspan'],
                'colspan':   mi['colspan'],
                'is_header': is_header,
                'bg':        bg,
                'is_merged_child': False,
            })
        grid_rows.append(row_cells)

    wb.close()

    return {
        'grid':            grid_rows,
        'num_header_rows': num_header_rows,
        'total_cols':      max_col,
        'source':          'excel',
    }


def read_submission_data(file_bytes: bytes) -> list:
    """
    Baca data submission Excel dan kembalikan sebagai list of dicts.

    Setiap baris data dikembalikan sebagai:
      { '__row_label': ..., '__col_0': ..., '__col_1': ..., ... }

    Kunci:
    - '__row_label'  → nilai kolom 1 (jika has_first_col)
    - '__col_N'      → nilai kolom ke-N dari kolom data (0-indexed)

    Baris yang SEPENUHNYA kosong di-skip (tidak ada data sama sekali).
    Baris dengan sebagian kosong TETAP diikutsertakan.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    max_row = ws.max_row or 1
    max_col = _get_actual_max_col(ws)

    top_left, non_tl = _build_merge_index(ws)
    num_header_rows  = detect_header_rows(ws)

    # Deteksi has_first_col
    has_first_col = False
    mi1 = top_left.get((1, 1))
    if mi1 and mi1['rowspan'] >= num_header_rows and num_header_rows > 1:
        has_first_col = True
    elif num_header_rows == 1:
        for r in range(num_header_rows + 1, max_row + 1):
            if _cell_str(ws.cell(r, 1)).strip():
                has_first_col = True
                break

    data_start   = num_header_rows + 1
    num_data_cols = max_col - (1 if has_first_col else 0)
    col_offset    = 2 if has_first_col else 1

    rows = []
    for r in range(data_start, max_row + 1):
        # Kumpulkan semua nilai baris ini
        vals = [_cell_str(ws.cell(r, c)) for c in range(1, max_col + 1)]
        # Skip baris yang benar-benar kosong (semua '' atau None)
        if not any(v.strip() for v in vals):
            continue

        row_dict = {}
        if has_first_col:
            row_dict['__row_label'] = _cell_str(ws.cell(r, 1))
        for ci in range(num_data_cols):
            col = col_offset + ci
            if col <= max_col:
                row_dict[f'__col_{ci}'] = _cell_str(ws.cell(r, col))
            else:
                row_dict[f'__col_{ci}'] = ''
        rows.append(row_dict)

    wb.close()
    return rows


def parse_excel_schema_from_template(file_bytes: bytes) -> dict:
    """
    Baca struktur template Excel dan kembalikan sebagai fields_schema.

    PRINSIP: Tidak menebak atau merekonstruksi struktur.
    Menyimpan informasi posisi, merged ranges, dan label PERSIS dari file.

    Output format (fields_schema):
    {
      "header_levels": [
        // Level 0..N-1: setiap level = list sel header di baris tersebut
        // Setiap item: {label, colspan, rowspan, col, row}
        [{label, colspan, rowspan, col, row}, ...]
      ],
      "first_column": {
        "enabled": bool,
        "label": str,
        "col": int,           # nomor kolom (1-indexed)
        "default_rows": [str] # nilai dari baris data di kolom ini
      },
      "raw_headers": [
        // Representasi mentah setiap sel header: posisi + merge info
        {row, col, value, rowspan, colspan, is_full_rowspan}
      ],
      "num_header_rows": int,
      "total_cols": int,
    }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    max_row = ws.max_row or 1
    max_col = _get_actual_max_col(ws)

    top_left, non_tl = _build_merge_index(ws)
    num_header_rows  = detect_header_rows(ws)

    # ── raw_headers: semua sel header dengan metadata penuh ──────────────────
    raw_headers = []
    for r in range(1, num_header_rows + 1):
        for c in range(1, max_col + 1):
            if (r, c) in non_tl:
                continue
            cell = ws.cell(r, c)
            mi   = top_left.get((r, c), {'rowspan': 1, 'colspan': 1,
                                          'max_r': r, 'max_c': c})
            is_full_rowspan = (mi['rowspan'] >= num_header_rows and num_header_rows > 1)
            raw_headers.append({
                'row':            r,
                'col':            c,
                'value':          _cell_str(cell),
                'rowspan':        mi['rowspan'],
                'colspan':        mi['colspan'],
                'is_full_rowspan': is_full_rowspan,
            })

    # ── Deteksi first_column ──────────────────────────────────────────────────
    first_column = {'enabled': False, 'label': '', 'col': 1, 'default_rows': []}
    mi1 = top_left.get((1, 1))
    if mi1 and mi1['rowspan'] >= num_header_rows and num_header_rows > 1:
        # Multi-level: kolom 1 di-merge vertikal → first_column
        first_column['enabled'] = True
        first_column['label']   = _cell_str(ws.cell(1, 1))
        first_column['col']     = 1
    elif num_header_rows == 1:
        # Single-header: scan kolom 1 di data rows
        data_start = 2
        fc_vals = []
        for r in range(data_start, max_row + 1):
            v = _cell_str(ws.cell(r, 1)).strip()
            if v:
                fc_vals.append(v)
        if fc_vals:
            first_column['enabled']      = True
            first_column['label']        = _cell_str(ws.cell(1, 1))
            first_column['col']          = 1
            first_column['default_rows'] = fc_vals[:50]

    if first_column['enabled'] and not first_column['default_rows']:
        # Isi default_rows dari baris data
        data_start = num_header_rows + 1
        fc_vals = []
        for r in range(data_start, max_row + 1):
            v = _cell_str(ws.cell(r, first_column['col'])).strip()
            fc_vals.append(v)
        first_column['default_rows'] = fc_vals[:50]

    # ── header_levels: bangun per baris header ────────────────────────────────
    # Setiap level = sel-sel yang bukan first_column
    fc_col = first_column['col'] if first_column['enabled'] else None
    header_levels = []
    for r in range(1, num_header_rows + 1):
        level = []
        for h in raw_headers:
            if h['row'] != r:
                continue
            if fc_col and h['col'] == fc_col:
                continue
            level.append({
                'label':   h['value'],
                'name':    h['value'].lower().replace(' ', '_').replace('\n', '_')[:40] or f'col_{h["col"]}',
                'colspan': h['colspan'],
                'rowspan': h['rowspan'],
                'col':     h['col'],
                'is_leaf': (h['row'] == num_header_rows) or
                           (h['rowspan'] + h['row'] - 1 >= num_header_rows),
            })
        if level:
            header_levels.append(level)

    wb.close()

    return {
        'header_levels':   header_levels,
        'first_column':    first_column,
        'raw_headers':     raw_headers,
        'num_header_rows': num_header_rows,
        'total_cols':      max_col,
    }
