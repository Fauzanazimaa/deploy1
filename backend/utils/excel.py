"""
Excel utilities — parse structure & generate template.

fields_schema format (JSON):
{
  "header_levels": [
    // Level 0..N-2: group headers dengan span
    [{"label": "Kelompok A", "span": 2}, {"label": "Kelompok B", "span": 3}],
    // Level terakhir: kolom data sesungguhnya (leaf)
    [{"label": "Kol 1", "name": "kol_1"},
     {"label": "Kol 2", "name": "kol_2"}, ...]
  ],
  "first_column": {
    "enabled": true,
    "label": "Kecamatan",
    "default_rows": ["Kec. A", "Kec. B"]
  }
}
"""

import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styles ────────────────────────────────────────────────────────────────────

_COLORS = ['1E3A5F', '2563EB', '3B82F6', '60A5FA']
_thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                 top=Side(style='thin'),  bottom=Side(style='thin'))
_halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
_valign = Alignment(vertical='center')

def _hfont(lvl=0): return Font(bold=True, color='FFFFFF', size=11 if lvl==0 else 10)
def _hfill(lvl=0):
    c = _COLORS[min(lvl, len(_COLORS)-1)]
    return PatternFill(start_color=c, end_color=c, fill_type='solid')


# ── Schema normalize ──────────────────────────────────────────────────────────

def normalize_schema(raw) -> dict:
    """Normalisasi berbagai format schema ke format baru."""
    empty_fc = {'enabled': False, 'label': '', 'default_rows': []}
    if isinstance(raw, list):
        return {'header_levels': [raw], 'first_column': empty_fc}
    if isinstance(raw, dict):
        s = dict(raw)
        s.setdefault('first_column', empty_fc)
        s.setdefault('header_levels', [[]])
        return s
    return {'header_levels': [[]], 'first_column': empty_fc}


def get_leaf_columns(schema: dict) -> list:
    levels = schema.get('header_levels', [[]])
    return levels[-1] if levels else []


def _to_field_name(label: str) -> str:
    name = label.lower().strip()
    name = re.sub(r'[^a-z0-9]+', '_', name)
    name = name.strip('_') or 'kolom'
    return name


# ── Parse Excel ───────────────────────────────────────────────────────────────

def parse_excel_structure(file_bytes: bytes) -> dict:
    """
    Baca file Excel dan deteksi struktur header secara akurat.
    - Mendukung merged cells (horizontal dan vertikal)
    - Mendukung 1, 2, 3+ level header
    - Kolom pertama dideteksi otomatis jika ada isian baris konsisten
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    # ── Bangun grid nilai dengan resolusi merged cells ────────────────────────
    # grid[row][col] = nilai asli (dari top-left merged range)
    grid = {}
    for r in range(1, max_row + 1):
        grid[r] = {}
        for c in range(1, max_col + 1):
            grid[r][c] = ws.cell(r, c).value

    # Info merge: (r,c) -> MergedCellRange object (hanya top-left yang punya value asli)
    merge_map = {}   # (r,c) -> {'min_r','min_c','max_r','max_c'}
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[(r, c)] = {
                    'min_r': mr.min_row, 'min_c': mr.min_col,
                    'max_r': mr.max_row, 'max_c': mr.max_col,
                }
                # Semua sel dalam range dapat nilai dari top-left
                if (r, c) != (mr.min_row, mr.min_col):
                    grid[r][c] = grid[mr.min_row][mr.min_col]

    # ── Tentukan jumlah baris header ──────────────────────────────────────────
    # Cari merge vertikal (row_span > 1) → jumlah header = max row dari semua merge
    max_header_row = 1
    for (r, c), m in merge_map.items():
        if m['min_r'] == r and m['min_c'] == c:  # hanya proses top-left
            if m['max_r'] > max_header_row:
                max_header_row = m['max_r']
            if max_header_row >= 5:
                break

    # Verifikasi: baris setelah max_header_row harus punya data atau kosong
    # Jika semua sel di baris max_header_row+1 adalah None, kemungkinan bukan header
    num_header_rows = max_header_row

    # ── Bangun header_levels ──────────────────────────────────────────────────
    # Kita akan membangun per kolom terlebih dahulu, lalu konversi ke per level

    # Untuk setiap kolom data, tentukan path dari level 1 sampai level terakhir
    # col_path[c] = list of (label, row_span, col_span) per level
    
    # Pendekatan: untuk setiap sel di baris header, tentukan apakah itu group atau leaf
    
    if num_header_rows == 1:
        # ── Single level header ───────────────────────────────────────────────
        leaf = []
        for c in range(1, max_col + 1):
            val = grid[1].get(c)
            if val is None:
                mi = merge_map.get((1, c))
                if mi and mi['min_c'] != c:
                    continue
                continue
            mi = merge_map.get((1, c))
            if mi and mi['min_c'] != c:
                continue
            label = str(val).replace('*', '').strip()
            name  = _to_field_name(label)
            leaf.append({'label': label, 'name': name, 'type': 'text', 'required': False})
        header_levels = [leaf]

    else:
        # ── Multi level header ────────────────────────────────────────────────
        #
        # PENTING: Deteksi kolom yang di-rowspan penuh (dari baris 1 s/d num_header_rows)
        # → kolom ini bukan group header, melainkan "first_column" atau kolom tetap vertikal.
        # Kita perlu tahu kolom mana saja yang di-rowspan penuh agar tidak dimasukkan
        # ke header_levels sebagai group, melainkan ke first_column.
        #
        # Strategi: scan kolom 1 — jika punya merge vertikal dari row 1 s/d num_header_rows,
        # itu adalah first_column. Kolom itu di-skip dari semua level processing.

        # Cari kolom yang full-rowspan di baris 1
        full_rowspan_cols = set()
        for c in range(1, max_col + 1):
            mi = merge_map.get((1, c))
            if mi and mi['min_r'] == 1 and mi['max_r'] >= num_header_rows and mi['min_c'] == c:
                # Ini kolom yang di-rowspan penuh dari baris 1 sampai baris terakhir header
                # Tandai semua kolom dalam range ini (biasanya hanya 1 kolom jika col_span=1)
                for cc in range(mi['min_c'], mi['max_c'] + 1):
                    full_rowspan_cols.add(cc)

        header_levels = []

        for lvl_row in range(1, num_header_rows + 1):
            level_items = []
            processed = set()

            for c in range(1, max_col + 1):
                if c in processed:
                    continue

                # Skip kolom yang full-rowspan (akan jadi first_column)
                if c in full_rowspan_cols:
                    processed.add(c)
                    continue

                val = grid[lvl_row].get(c)
                mi  = merge_map.get((lvl_row, c))

                if mi:
                    # Hanya proses top-left dari range
                    if mi['min_r'] != lvl_row or mi['min_c'] != c:
                        processed.add(c)
                        continue

                    label    = str(val).strip() if val is not None else ''
                    col_span = mi['max_c'] - mi['min_c'] + 1
                    row_span = mi['max_r'] - mi['min_r'] + 1

                    # Kurangi col_span untuk kolom yang sudah jadi first_column
                    effective_span = sum(
                        1 for cc in range(mi['min_c'], mi['max_c'] + 1)
                        if cc not in full_rowspan_cols
                    )

                    for cc in range(c, c + col_span):
                        processed.add(cc)

                    if row_span > 1 and (lvl_row + row_span - 1) >= num_header_rows:
                        # Merge vertikal sampai baris terakhir header → leaf
                        name = _to_field_name(label)
                        level_items.append({'label': label, 'name': name,
                                            'type': 'text', 'required': False})
                    elif lvl_row < num_header_rows:
                        # Merge horizontal di baris non-terakhir → group header
                        if effective_span > 0:
                            level_items.append({'label': label, 'span': effective_span})
                    else:
                        # Baris terakhir header → leaf
                        name = _to_field_name(label)
                        level_items.append({'label': label, 'name': name,
                                            'type': 'text', 'required': False})
                else:
                    # Sel biasa (tidak di-merge)
                    processed.add(c)
                    if val is None:
                        continue
                    label = str(val).replace('*', '').strip()
                    if not label:
                        continue
                    required = str(val).strip().endswith('*')

                    if lvl_row < num_header_rows:
                        level_items.append({'label': label, 'span': 1})
                    else:
                        name = _to_field_name(label)
                        level_items.append({'label': label, 'name': name,
                                            'type': 'text', 'required': required})

            if level_items:
                header_levels.append(level_items)

        # Ekstrak first_column dari full_rowspan_cols (kolom pertama saja)
        if full_rowspan_cols:
            fc_col = min(full_rowspan_cols)  # ambil kolom terkiri
            fc_val = grid[1].get(fc_col)
            if fc_val is not None:
                # Override first_column label dari nilai sel rowspan
                _fc_label_override = str(fc_val).strip()
            else:
                _fc_label_override = None
        else:
            _fc_label_override = None

    # ── Deteksi kolom pertama ─────────────────────────────────────────────────
    # Cek baris data: apakah kolom 1 berisi teks (nama kecamatan dll)?
    data_start   = num_header_rows + 1
    first_col_vals = []
    for r in range(data_start, min(data_start + 30, max_row + 1)):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip():
            first_col_vals.append(str(v).strip())

    # Label kolom pertama:
    # 1. Jika ada full_rowspan_cols (multi-level), gunakan label dari sel rowspan
    # 2. Jika tidak, ambil dari level terakhir kolom pertama
    if num_header_rows > 1 and '_fc_label_override' in dir() and _fc_label_override:
        first_col_label = _fc_label_override
        has_first = True  # Ada full_rowspan → pasti first_column
    else:
        # Single level atau tidak ada rowspan
        first_col_label = ''
        if header_levels:
            last = header_levels[-1]
            if last:
                first_col_label = last[0].get('label', '')

        has_first = len(first_col_vals) > 0

        # Jika kolom pertama ada isian, hapus dari leaf columns
        if has_first and header_levels:
            last = header_levels[-1]
            if last and last[0].get('name'):
                header_levels[-1] = last[1:]

    first_column = {
        'enabled': has_first,
        'label':   first_col_label,
        'default_rows': first_col_vals[:50],
    }

    wb.close()
    return {
        'header_levels': header_levels,
        'first_column':  first_column,
    }


# ── Generate template ─────────────────────────────────────────────────────────

def generate_template(data_type) -> bytes:
    raw    = data_type.get_fields_schema()
    schema = normalize_schema(raw)
    return _build_excel(data_type.name, schema)


def _build_excel(sheet_name: str, schema: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    levels       = schema.get('header_levels', [[]])
    first_col    = schema.get('first_column', {})
    has_first    = first_col.get('enabled', False)
    first_label  = first_col.get('label', '')
    default_rows = first_col.get('default_rows', [])

    num_levels    = len(levels)
    leaf_cols     = levels[-1] if levels else []
    num_data_cols = len(leaf_cols)
    col_offset    = 2 if has_first else 1

    # ── Header rows ───────────────────────────────────────────────────────────
    for lvl_idx, level in enumerate(levels):
        row = lvl_idx + 1

        # Kolom pertama (merge semua baris header)
        if has_first:
            cell = ws.cell(row=row, column=1)
            if lvl_idx == 0:
                if num_levels > 1:
                    ws.merge_cells(start_row=1, start_column=1,
                                   end_row=num_levels, end_column=1)
                cell.value     = first_label
                cell.font      = Font(bold=True, color='FFFFFF', size=11)
                cell.fill      = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
                cell.alignment = _halign
                cell.border    = _thin

        is_last = (lvl_idx == num_levels - 1)

        if not is_last:
            # Group header dengan span
            col = col_offset
            for grp in level:
                span  = grp.get('span', 1)
                label = grp.get('label', '')
                cell  = ws.cell(row=row, column=col)
                cell.value     = label
                cell.font      = _hfont(lvl_idx)
                cell.fill      = _hfill(lvl_idx)
                cell.alignment = _halign
                cell.border    = _thin
                if span > 1:
                    ws.merge_cells(start_row=row, start_column=col,
                                   end_row=row, end_column=col + span - 1)
                    for cc in range(col, col + span):
                        ws.cell(row=row, column=cc).border = _thin
                col += span
        else:
            # Leaf columns
            for ci, field in enumerate(level):
                col   = col_offset + ci
                label = field.get('label', field.get('name', f'Kolom {ci+1}'))
                req   = field.get('required', False)
                cell  = ws.cell(row=row, column=col)
                cell.value     = f"{label}{'*' if req else ''}"
                cell.font      = _hfont(lvl_idx)
                cell.fill      = _hfill(lvl_idx)
                cell.alignment = _halign
                cell.border    = _thin
                ws.column_dimensions[get_column_letter(col)].width = max(15, len(label) + 5)

    if has_first:
        ws.column_dimensions['A'].width = max(20, len(first_label) + 5)

    for r in range(1, num_levels + 1):
        ws.row_dimensions[r].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    data_start = num_levels + 1
    n_rows     = max(len(default_rows), 10)

    _fc_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    _fc_font = Font(bold=True, size=10)

    for ri in range(n_rows):
        row = data_start + ri
        if has_first:
            cell = ws.cell(row=row, column=1)
            cell.value     = default_rows[ri] if ri < len(default_rows) else ''
            cell.border    = _thin
            cell.alignment = _valign
            if ri < len(default_rows):
                cell.fill = _fc_fill
                cell.font = _fc_font
        for ci in range(num_data_cols):
            cell = ws.cell(row=row, column=col_offset + ci)
            cell.border    = _thin
            cell.alignment = _valign

    ws.freeze_panes = f"{get_column_letter(col_offset)}{data_start}"

    # Note
    note = ws.cell(row=data_start + n_rows + 1, column=1)
    note.value = '* Kolom wajib diisi. Jangan ubah baris header.'
    note.font  = Font(italic=True, color='888888', size=9)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ── Read data ─────────────────────────────────────────────────────────────────

def read_excel_data(file_bytes: bytes, fields_schema) -> list:
    schema     = normalize_schema(fields_schema)
    leaf_cols  = get_leaf_columns(schema)
    first_col  = schema.get('first_column', {})
    has_first  = first_col.get('enabled', False)
    num_levels = len(schema.get('header_levels', [[]]))

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    data_start = num_levels + 1
    col_offset = 2 if has_first else 1

    col_to_field = {col_offset + ci: f.get('name', f'col_{ci}')
                    for ci, f in enumerate(leaf_cols)}

    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if all(v is None for v in row):
            continue
        rd = {}
        if has_first and row:
            key = _to_field_name(first_col.get('label', 'kolom_1'))
            rd[key] = row[0]
        for ci, val in enumerate(row, start=1):
            if ci in col_to_field:
                rd[col_to_field[ci]] = val
        if any(v is not None for v in rd.values()):
            rows.append(rd)

    wb.close()
    return rows
